"""
Import historical Sales & COGs data from Excel (.xlsx) or CSV files.
Supports both the OneOps template format and the SkyT2 CSV format.
Also supports multi-store GoMenu CSV files (one CSV containing data for
multiple sub-stores under a Company).
"""
import io
import csv
import re
import calendar
import openpyxl
from decimal import Decimal, InvalidOperation
from datetime import date

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser

from users.views import get_target_org
from users.models import Organization
from closing.models import (
    DailyClosing, ClosingOtherSale, ClosingSupplierCost,
    ClosingCashExpense, ClosingHRCash,
    Supplier, SalesCategory,
)

# These names are skipped (totals, variance, headers)
SKIP_NAMES = {'income total', 'cogs total', 'total', 'variance', 'income', 'store name', 'month'}

SURCHARGE_NAME = 'surcharge'


def _safe_decimal(val):
    """Convert cell value to Decimal, return None if empty/zero/invalid."""
    if val is None or val == '' or val == 0:
        return None
    try:
        d = Decimal(str(val).strip()).quantize(Decimal('0.01'))
        return d if d != 0 else None
    except (InvalidOperation, ValueError):
        return None


def _process_day_data(org, closing_date, income_data, cogs_data, result):
    """
    Process one day's income and COGs data.
    income_data: list of (name, value) tuples
    cogs_data: list of (name, value) tuples
    """
    # Check if any data exists
    has_data = any(v is not None for _, v in income_data + cogs_data)
    if not has_data:
        return

    closing, created = DailyClosing.objects.get_or_create(
        organization=org,
        closing_date=closing_date,
        defaults={'status': 'APPROVED'}
    )
    if created:
        result['closings_created'] += 1
    else:
        result['closings_updated'] += 1
        closing.other_sales.all().delete()
        closing.supplier_costs.all().delete()

    pos_cash = Decimal('0')
    pos_card = Decimal('0')
    surcharge = Decimal('0')

    for name, val in income_data:
        if val is None:
            continue
        name_lower = name.lower().strip()
        if name_lower == 'cash':
            pos_cash = val
        elif name_lower == 'card':
            pos_card = val
        elif name_lower == SURCHARGE_NAME:
            surcharge = val
        else:
            ClosingOtherSale.objects.create(closing=closing, name=name.strip(), amount=val)
            result['other_sales_created'] += 1
            _, sc_created = SalesCategory.objects.get_or_create(
                organization=org, name=name.strip(),
                defaults={'is_active': True}
            )
            if sc_created and name.strip() not in result['sales_categories_added']:
                result['sales_categories_added'].append(name.strip())

    closing.pos_cash = pos_cash
    closing.pos_card = pos_card + surcharge
    closing.actual_cash = pos_cash
    closing.actual_card = pos_card + surcharge
    closing.save()

    for name, val in cogs_data:
        if val is None:
            continue
        name_clean = name.strip()
        name_lower = name_clean.lower()
        if name_lower == 'other':
            supplier, _ = Supplier.objects.get_or_create(
                organization=org, code='OTHER',
                defaults={'name': 'Other', 'category': 'COGS'}
            )
        else:
            code = name_clean.upper().replace(' ', '_')[:50]
            supplier, sup_created = Supplier.objects.get_or_create(
                organization=org, code=code,
                defaults={'name': name_clean, 'category': 'COGS'}
            )
            if sup_created and name_clean not in result['suppliers_added']:
                result['suppliers_added'].append(name_clean)

        ClosingSupplierCost.objects.create(closing=closing, supplier=supplier, amount=val)
        result['supplier_costs_created'] += 1


def _new_result():
    return {
        'closings_created': 0, 'closings_updated': 0,
        'other_sales_created': 0, 'supplier_costs_created': 0,
        'suppliers_added': [], 'sales_categories_added': [],
    }


def _merge_result(stats, result):
    stats['closings_created'] += result['closings_created']
    stats['closings_updated'] += result['closings_updated']
    stats['other_sales_created'] += result['other_sales_created']
    stats['supplier_costs_created'] += result['supplier_costs_created']
    for s in result.get('suppliers_added', []):
        if s not in stats['suppliers_added']:
            stats['suppliers_added'].append(s)
    for s in result.get('sales_categories_added', []):
        if s not in stats['sales_categories_added']:
            stats['sales_categories_added'].append(s)


class ImportDataView(APIView):
    """Upload Excel or CSV file to import historical sales & COGs data."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        org = get_target_org(request)
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file uploaded'}, status=400)

        filename = file.name.lower()
        stats = {
            'months_processed': 0,
            'closings_created': 0,
            'closings_updated': 0,
            'other_sales_created': 0,
            'supplier_costs_created': 0,
            'suppliers_added': [],
            'sales_categories_added': [],
            'errors': [],
        }

        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            # Detect format from headers
            if rows and len(rows[0]) == 2 and rows[0][0].strip().lower() == 'date':
                self._process_simple_csv(rows, org, stats)
            elif rows and any('Income -' in h or 'COGs -' in h or 'Cash Tracking' in h for h in rows[0]):
                self._process_gomenu_csv(rows, org, stats, request.user)
            else:
                self._process_csv_from_rows(rows, org, stats)
        elif filename.endswith(('.xlsx', '.xls')):
            self._process_excel(file, org, stats)
        else:
            return Response({'error': 'Unsupported file format. Use .xlsx or .csv'}, status=400)

        return Response(stats)

    def _process_simple_csv(self, rows, org, stats):
        """
        Process simple CSV format:
        Date,Daily Total
        2024-01-01,3913.51
        """
        months_seen = set()
        for row in rows[1:]:  # skip header
            if len(row) < 2 or not row[0].strip():
                continue
            try:
                raw = row[0].strip()
                if '/' in raw:
                    parts = raw.split('/')
                    d = date(int(parts[2]), int(parts[1]), int(parts[0])) if len(parts) == 3 else date.fromisoformat(raw)
                else:
                    d = date.fromisoformat(raw)
                amount = _safe_decimal(row[1])
                if amount is None:
                    continue
                closing, created = DailyClosing.objects.get_or_create(
                    organization=org,
                    closing_date=d,
                    defaults={'status': 'APPROVED'}
                )
                closing.pos_cash = amount
                closing.pos_card = Decimal('0')
                closing.actual_cash = amount
                closing.actual_card = Decimal('0')
                closing.other_sales.all().delete()
                closing.supplier_costs.all().delete()
                closing.save()
                month_key = (d.year, d.month)
                months_seen.add(month_key)
                if created:
                    stats['closings_created'] += 1
                else:
                    stats['closings_updated'] += 1
            except (ValueError, Exception) as e:
                stats['errors'].append(f'Row {row[0]}: {str(e)}')
        stats['months_processed'] = len(months_seen)

    def _detect_csv_store_names(self, headers):
        """
        Detect unique store names from GoMenu CSV headers.
        Looks for patterns like "Income - STORE_NAME | Total".
        Returns a list of unique store name strings found in the CSV.
        """
        store_names = []
        for h in headers:
            h = h.strip().strip('"')
            # Match "Income - STORE_NAME | Total"
            m = re.match(r'^Income\s*-\s*(.+?)\s*\|\s*Total$', h, re.IGNORECASE)
            if m:
                name = m.group(1).strip()
                if name not in store_names:
                    store_names.append(name)
        return store_names

    def _fuzzy_match_store(self, csv_name, sub_stores):
        """
        Try to match a CSV store name to a sub-store Organization object.
        Uses case-insensitive partial matching.
        Returns the matched Organization or None.
        """
        csv_lower = csv_name.lower().strip()
        # Also try without leading "The "
        csv_no_the = re.sub(r'^the\s+', '', csv_lower)

        for sub in sub_stores:
            db_lower = sub.name.lower().strip()
            db_no_the = re.sub(r'^the\s+', '', db_lower)

            # Exact match
            if csv_lower == db_lower:
                return sub
            # CSV name contains DB name or vice versa
            if db_lower in csv_lower or csv_lower in db_lower:
                return sub
            # Try without "The " prefix
            if db_no_the in csv_no_the or csv_no_the in db_no_the:
                return sub
            # Try without "The " on just one side
            if db_no_the in csv_lower or csv_lower in db_no_the:
                return sub
            if db_lower in csv_no_the or csv_no_the in db_lower:
                return sub
        return None

    def _extract_store_name_from_header(self, header):
        """
        Extract store name from a header like "Income - Store Name | Field"
        or "COGs - Store Name | Field" or "Expenses - Store Name | Field".
        Returns (store_name, field_name) or (None, None) if not matched.
        """
        header = header.strip().strip('"')
        m = re.match(r'^(Income|COGs|Expenses)\s*-\s*(.+?)\s*\|\s*(.+)$', header, re.IGNORECASE)
        if m:
            return m.group(2).strip(), m.group(3).strip()
        return None, None

    def _process_gomenu_csv(self, rows, org, stats, user):
        """
        Process GoMenu POS export CSV format.
        Supports multi-store CSVs: if the CSV contains data for multiple stores
        and the target org is a Company (has sub_stores), each store's data is
        saved to the matching sub-store, and Cash Tracking goes to the parent.

        Headers like: Date, Income - Store | Cash, Income - Store | Card, ...,
                      COGs - Store | Supplier, ..., Cash Tracking | Cash, ...
        """
        headers = rows[0]
        # Strip quotes from headers
        headers = [h.strip().strip('"') for h in headers]

        # Detect store names in CSV
        csv_store_names = self._detect_csv_store_names(headers)
        is_multi_store = len(csv_store_names) > 1 and org.is_company
        sub_stores = list(org.sub_stores.all()) if is_multi_store else []

        # Build store mapping: csv_store_name -> Organization
        store_map = {}  # csv_name -> Organization
        unmatched_stores = []
        if is_multi_store:
            for csv_name in csv_store_names:
                matched = self._fuzzy_match_store(csv_name, sub_stores)
                if matched:
                    store_map[csv_name] = matched
                else:
                    unmatched_stores.append(csv_name)
            if unmatched_stores:
                stats['errors'].append(
                    f'Could not match CSV stores to sub-stores: {", ".join(unmatched_stores)}'
                )
            stats['multi_store'] = True
            stats['store_mapping'] = {
                csv_name: {'id': s.id, 'name': s.name}
                for csv_name, s in store_map.items()
            }
            # Per-store stats
            stats['per_store'] = {}
            for csv_name, s in store_map.items():
                stats['per_store'][s.name] = {'closings_created': 0, 'id': s.id}
            stats['per_store'][org.name + ' (Cash Tracking)'] = {'closings_created': 0, 'id': org.id}

        # Parse column mappings from headers
        # For multi-store: col_map entries include store_name
        # col_map: list of (index, section, field_name, store_csv_name)
        col_map = []
        for i, h in enumerate(headers):
            h_clean = h.strip()
            if h_clean.lower() == 'date':
                col_map.append((i, 'date', None, None))
                continue

            if '|' not in h_clean:
                col_map.append((i, 'skip', None, None))
                continue

            parts = h_clean.split('|')
            field_name = parts[-1].strip()
            prefix = parts[0].strip()

            # Detect store name from prefix
            store_csv_name = None
            prefix_lower = prefix.lower()

            if is_multi_store:
                # Extract store name from "Income - StoreName" etc.
                sn, fn = self._extract_store_name_from_header(h_clean)
                if sn:
                    store_csv_name = sn
                    field_name = fn

            if 'income' in prefix_lower:
                fl = field_name.lower()
                if fl == 'total':
                    col_map.append((i, 'income_total', None, store_csv_name))
                elif fl == 'cash':
                    col_map.append((i, 'income_cash', None, store_csv_name))
                elif fl == 'card':
                    col_map.append((i, 'income_card', None, store_csv_name))
                elif fl == 'surcharge':
                    col_map.append((i, 'income_surcharge', None, store_csv_name))
                else:
                    col_map.append((i, 'other_sale', field_name, store_csv_name))
            elif 'cogs' in prefix_lower:
                fl = field_name.lower()
                if fl == 'total':
                    col_map.append((i, 'cogs_total', None, store_csv_name))
                else:
                    col_map.append((i, 'cogs', field_name, store_csv_name))
            elif 'expense' in prefix_lower:
                fl = field_name.lower()
                if fl == 'total':
                    col_map.append((i, 'expenses_total', None, store_csv_name))
                else:
                    col_map.append((i, 'skip', None, store_csv_name))
            elif 'cash tracking' in prefix_lower:
                fl = field_name.lower()
                if fl == 'cash':
                    col_map.append((i, 'ct_cash', None, None))
                elif fl == 'banking':
                    col_map.append((i, 'ct_banking', None, None))
                elif fl == 'deposit':
                    col_map.append((i, 'ct_deposit', None, None))
                elif fl == 'cash movement':
                    col_map.append((i, 'ct_movement', None, None))
                else:
                    col_map.append((i, 'skip', None, None))
            else:
                col_map.append((i, 'skip', None, None))

        # First pass: collect all dates to determine range
        all_dates = []
        for row in rows[1:]:
            if not row or not row[0].strip():
                continue
            try:
                raw = row[0].strip().strip('"')
                if '/' in raw:
                    parts = raw.split('/')
                    if len(parts) == 3:
                        all_dates.append(date(int(parts[2]), int(parts[1]), int(parts[0])))
                else:
                    all_dates.append(date.fromisoformat(raw))
            except (ValueError, IndexError):
                continue

        # Delete ALL existing closings in the CSV date range (overwrite)
        if all_dates:
            min_date, max_date = min(all_dates), max(all_dates)
            if is_multi_store:
                # Delete for parent AND all matched sub-stores
                affected_org_ids = [org.id] + [s.id for s in store_map.values()]
                existing = DailyClosing.objects.filter(
                    organization_id__in=affected_org_ids,
                    closing_date__gte=min_date,
                    closing_date__lte=max_date,
                )
            else:
                existing = DailyClosing.objects.filter(
                    organization=org,
                    closing_date__gte=min_date,
                    closing_date__lte=max_date,
                )
            del_count = existing.count()
            existing.delete()
            stats['closings_deleted'] = del_count

        months_seen = set()

        for row in rows[1:]:
            if not row or not row[0].strip():
                continue
            try:
                raw = row[0].strip().strip('"')
                if '/' in raw:
                    parts = raw.split('/')
                    if len(parts) == 3:
                        d = date(int(parts[2]), int(parts[1]), int(parts[0]))  # DD/MM/YYYY
                    else:
                        continue
                else:
                    d = date.fromisoformat(raw)
            except (ValueError, IndexError):
                continue

            try:
                if is_multi_store:
                    self._process_gomenu_row_multi(
                        row, d, col_map, org, store_map, stats, user, months_seen
                    )
                else:
                    self._process_gomenu_row_single(
                        row, d, col_map, org, stats, user, months_seen
                    )
            except Exception as e:
                stats['errors'].append(f'{d}: {str(e)}')

        stats['months_processed'] = len(months_seen)

    def _process_gomenu_row_single(self, row, d, col_map, org, stats, user, months_seen):
        """Process a single row of GoMenu CSV for single-store mode (original behavior)."""
        income_cash = Decimal('0')
        income_card = Decimal('0')
        surcharge = Decimal('0')
        other_sales = []
        cogs_items = []
        ct_cash = None
        ct_banking = None
        ct_deposit = None
        ct_movement = None

        for idx, section, field_name, _store in col_map:
            if section in ('date', 'skip', 'income_total', 'cogs_total', 'expenses_total'):
                continue
            val = _safe_decimal(row[idx]) if idx < len(row) else None
            if val is None:
                continue

            if section == 'income_cash':
                income_cash = val
            elif section == 'income_card':
                income_card = val
            elif section == 'income_surcharge':
                surcharge = val
            elif section == 'other_sale':
                other_sales.append((field_name, val))
            elif section == 'cogs':
                cogs_items.append((field_name, val))
            elif section == 'ct_cash':
                ct_cash = val
            elif section == 'ct_banking':
                ct_banking = val
            elif section == 'ct_deposit':
                ct_deposit = val
            elif section == 'ct_movement':
                ct_movement = val

        has_data = (income_cash or income_card or surcharge or
                    other_sales or cogs_items or
                    ct_cash or ct_banking or ct_deposit or ct_movement)
        if not has_data:
            return

        closing = DailyClosing.objects.create(
            organization=org,
            closing_date=d,
            status='APPROVED',
        )
        stats['closings_created'] += 1

        closing.pos_cash = income_cash
        closing.pos_card = income_card + surcharge
        closing.actual_cash = ct_cash if ct_cash is not None else income_cash
        closing.actual_card = income_card + surcharge
        closing.bank_deposit = ct_banking if ct_banking is not None else Decimal('0')
        closing.save()

        for name, val in other_sales:
            ClosingOtherSale.objects.create(closing=closing, name=name, amount=val)
            stats['other_sales_created'] += 1
            _, sc_created = SalesCategory.objects.get_or_create(
                organization=org, name=name,
                defaults={'is_active': True}
            )
            if sc_created and name not in stats['sales_categories_added']:
                stats['sales_categories_added'].append(name)

        for name, val in cogs_items:
            name_clean = name.strip()
            code = name_clean.upper().replace(' ', '_')[:50]
            if name_clean.lower() == 'other':
                supplier, _ = Supplier.objects.get_or_create(
                    organization=org, code='OTHER',
                    defaults={'name': 'Other', 'category': 'COGS'}
                )
            else:
                supplier, sup_created = Supplier.objects.get_or_create(
                    organization=org, code=code,
                    defaults={'name': name_clean, 'category': 'COGS'}
                )
                if sup_created and name_clean not in stats['suppliers_added']:
                    stats['suppliers_added'].append(name_clean)
            ClosingSupplierCost.objects.create(
                closing=closing, supplier=supplier, amount=val
            )
            stats['supplier_costs_created'] += 1

        if ct_deposit is not None and ct_deposit != 0:
            ClosingCashExpense.objects.create(
                daily_closing=closing,
                category='OTHER',
                reason='Deposit',
                amount=abs(ct_deposit),
                created_by=user,
            )

        months_seen.add((d.year, d.month))

    def _process_gomenu_row_multi(self, row, d, col_map, org, store_map, stats, user, months_seen):
        """
        Process a single row of GoMenu CSV for multi-store mode.
        Creates separate DailyClosing records for each sub-store and one for
        the parent company (Cash Tracking).
        """
        # Collect data per store
        # store_data[csv_store_name] = {income_cash, income_card, surcharge, other_sales, cogs_items, income_total, cogs_total, expenses_total}
        store_data = {}
        ct_cash = None
        ct_banking = None
        ct_deposit = None
        ct_movement = None

        for idx, section, field_name, store_csv_name in col_map:
            if section in ('date', 'skip'):
                continue
            val = _safe_decimal(row[idx]) if idx < len(row) else None

            # Cash Tracking columns (no store_csv_name)
            if section.startswith('ct_'):
                if val is None:
                    continue
                if section == 'ct_cash':
                    ct_cash = val
                elif section == 'ct_banking':
                    ct_banking = val
                elif section == 'ct_deposit':
                    ct_deposit = val
                elif section == 'ct_movement':
                    ct_movement = val
                continue

            if store_csv_name is None:
                continue

            # Initialize store data if needed
            if store_csv_name not in store_data:
                store_data[store_csv_name] = {
                    'income_cash': Decimal('0'),
                    'income_card': Decimal('0'),
                    'surcharge': Decimal('0'),
                    'other_sales': [],
                    'cogs_items': [],
                    'income_total': Decimal('0'),
                    'cogs_total': Decimal('0'),
                    'expenses_total': Decimal('0'),
                }

            sd = store_data[store_csv_name]

            if val is None:
                continue

            if section == 'income_cash':
                sd['income_cash'] = val
            elif section == 'income_card':
                sd['income_card'] = val
            elif section == 'income_surcharge':
                sd['surcharge'] = val
            elif section == 'income_total':
                sd['income_total'] = val
            elif section == 'other_sale':
                sd['other_sales'].append((field_name, val))
            elif section == 'cogs':
                sd['cogs_items'].append((field_name, val))
            elif section == 'cogs_total':
                sd['cogs_total'] = val
            elif section == 'expenses_total':
                sd['expenses_total'] = val

        # Check if there's any data at all
        has_any = any(
            sd['income_cash'] or sd['income_card'] or sd['surcharge'] or
            sd['other_sales'] or sd['cogs_items'] or sd['income_total']
            for sd in store_data.values()
        ) or ct_cash or ct_banking or ct_deposit or ct_movement
        if not has_any:
            return

        # Create DailyClosing for each matched sub-store
        for csv_name, sd in store_data.items():
            target_org = store_map.get(csv_name)
            if not target_org:
                continue  # unmatched store, skip

            has_store_data = (
                sd['income_cash'] or sd['income_card'] or sd['surcharge'] or
                sd['other_sales'] or sd['cogs_items'] or sd['income_total']
            )
            if not has_store_data:
                continue

            closing = DailyClosing.objects.create(
                organization=target_org,
                closing_date=d,
                status='APPROVED',
            )
            stats['closings_created'] += 1
            if target_org.name in stats.get('per_store', {}):
                stats['per_store'][target_org.name]['closings_created'] += 1

            # Income fields
            closing.pos_cash = sd['income_cash']
            closing.pos_card = sd['income_card'] + sd['surcharge']
            closing.actual_cash = sd['income_cash']
            closing.actual_card = sd['income_card'] + sd['surcharge']
            closing.save()

            # Other Sales
            for name, val in sd['other_sales']:
                ClosingOtherSale.objects.create(closing=closing, name=name, amount=val)
                stats['other_sales_created'] += 1
                _, sc_created = SalesCategory.objects.get_or_create(
                    organization=target_org, name=name,
                    defaults={'is_active': True}
                )
                if sc_created and name not in stats['sales_categories_added']:
                    stats['sales_categories_added'].append(name)

            # COGs → SupplierCosts
            for name, val in sd['cogs_items']:
                name_clean = name.strip()
                code = name_clean.upper().replace(' ', '_')[:50]
                if name_clean.lower() == 'other':
                    supplier, _ = Supplier.objects.get_or_create(
                        organization=target_org, code='OTHER',
                        defaults={'name': 'Other', 'category': 'COGS'}
                    )
                else:
                    supplier, sup_created = Supplier.objects.get_or_create(
                        organization=target_org, code=code,
                        defaults={'name': name_clean, 'category': 'COGS'}
                    )
                    if sup_created and name_clean not in stats['suppliers_added']:
                        stats['suppliers_added'].append(name_clean)
                ClosingSupplierCost.objects.create(
                    closing=closing, supplier=supplier, amount=val
                )
                stats['supplier_costs_created'] += 1

        # Cash Tracking → parent Company DailyClosing
        has_ct = ct_cash or ct_banking or ct_deposit or ct_movement
        if has_ct:
            ct_closing = DailyClosing.objects.create(
                organization=org,
                closing_date=d,
                status='APPROVED',
            )
            stats['closings_created'] += 1
            parent_key = org.name + ' (Cash Tracking)'
            if parent_key in stats.get('per_store', {}):
                stats['per_store'][parent_key]['closings_created'] += 1

            ct_closing.actual_cash = ct_cash if ct_cash is not None else Decimal('0')
            ct_closing.bank_deposit = ct_banking if ct_banking is not None else Decimal('0')
            ct_closing.save()

            if ct_deposit is not None and ct_deposit != 0:
                ClosingCashExpense.objects.create(
                    daily_closing=ct_closing,
                    category='OTHER',
                    reason='Deposit',
                    amount=abs(ct_deposit),
                    created_by=user,
                )

        months_seen.add((d.year, d.month))

    def _process_csv_from_rows(self, rows, org, stats):
        """
        Process CSV in SkyT2 format:
        Store Name,Sky T2
        Month,2024-01
        (blank)
        Income,Total,1,2,...,31
        Cash,total,val,val,...
        ...
        Income Total,...
        (blank)
        COGs (Supplier Costs),Total,1,2,...,31
        Coke,total,val,...
        ...
        COGs Total,...
        (repeats for each month)
        """
        month_blocks = []
        current_block = {'year': None, 'month': None, 'income': [], 'cogs': []}
        current_section = None

        for row in rows:
            if not row or all(c.strip() == '' for c in row):
                continue

            first = row[0].strip().lower() if row[0] else ''

            # Detect month line
            if first == 'month' and len(row) >= 2:
                # Save previous block if it has data
                if current_block['year'] and (current_block['income'] or current_block['cogs']):
                    month_blocks.append(current_block)
                # Parse YYYY-MM
                try:
                    parts = row[1].strip().split('-')
                    current_block = {
                        'year': int(parts[0]),
                        'month': int(parts[1]),
                        'income': [],
                        'cogs': [],
                    }
                    current_section = None
                except (ValueError, IndexError):
                    current_block = {'year': None, 'month': None, 'income': [], 'cogs': []}
                continue

            if first == 'store name':
                continue

            # Detect section headers
            if first == 'income':
                current_section = 'income'
                continue
            elif first.startswith('cogs'):
                current_section = 'cogs'
                continue

            # Skip totals/variance
            if first in SKIP_NAMES:
                current_section = None  # End of section
                continue

            # Data rows
            if current_block['year'] and current_section and first:
                name = row[0].strip()
                # Values start at index 2 (skip name and total)
                values = row[2:] if len(row) > 2 else []
                if current_section == 'income':
                    current_block['income'].append((name, values))
                elif current_section == 'cogs':
                    current_block['cogs'].append((name, values))

        # Don't forget the last block
        if current_block['year'] and (current_block['income'] or current_block['cogs']):
            month_blocks.append(current_block)

        # Process each month block
        for block in month_blocks:
            year, month = block['year'], block['month']
            days_in_month = calendar.monthrange(year, month)[1]
            result = _new_result()

            try:
                for day in range(1, days_in_month + 1):
                    day_idx = day - 1  # 0-based index into values

                    income_data = []
                    for name, values in block['income']:
                        val = _safe_decimal(values[day_idx]) if day_idx < len(values) else None
                        income_data.append((name, val))

                    cogs_data = []
                    for name, values in block['cogs']:
                        val = _safe_decimal(values[day_idx]) if day_idx < len(values) else None
                        cogs_data.append((name, val))

                    _process_day_data(org, date(year, month, day), income_data, cogs_data, result)

                stats['months_processed'] += 1
                _merge_result(stats, result)
            except Exception as e:
                stats['errors'].append(f'{year}-{month:02d}: {str(e)}')

    def _process_excel(self, file, org, stats):
        """Process Excel file (OneOps template format)."""
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            stats['errors'].append(f'Invalid Excel file: {str(e)}')
            return

        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == 'reference':
                continue
            try:
                parts = sheet_name.strip().split('-')
                year = int(parts[0])
                month = int(parts[1])
                if not (2020 <= year <= 2030 and 1 <= month <= 12):
                    continue
            except (ValueError, IndexError):
                continue

            ws = wb[sheet_name]
            days_in_month = calendar.monthrange(year, month)[1]
            result = _new_result()

            try:
                # Find sections
                income_rows = []
                cogs_rows = []
                current_section = None

                for row_idx in range(1, ws.max_row + 1):
                    cell_a = ws.cell(row=row_idx, column=1).value
                    if cell_a is None:
                        continue
                    cell_str = str(cell_a).strip().lower()
                    if cell_str == 'income':
                        current_section = 'income'
                        continue
                    elif cell_str.startswith('cogs'):
                        current_section = 'cogs'
                        continue
                    elif cell_str in SKIP_NAMES:
                        continue
                    if current_section == 'income' and cell_str:
                        income_rows.append((row_idx, str(cell_a).strip()))
                    elif current_section == 'cogs' and cell_str:
                        cogs_rows.append((row_idx, str(cell_a).strip()))

                for day in range(1, days_in_month + 1):
                    col = day + 2

                    income_data = []
                    for row_idx, name in income_rows:
                        val = _safe_decimal(ws.cell(row=row_idx, column=col).value)
                        income_data.append((name, val))

                    cogs_data = []
                    for row_idx, name in cogs_rows:
                        val = _safe_decimal(ws.cell(row=row_idx, column=col).value)
                        cogs_data.append((name, val))

                    _process_day_data(org, date(year, month, day), income_data, cogs_data, result)

                stats['months_processed'] += 1
                _merge_result(stats, result)
            except Exception as e:
                stats['errors'].append(f'{sheet_name}: {str(e)}')


class ImportTemplateView(APIView):
    """Download the import template Excel file."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        org = get_target_org(request)
        fmt = request.query_params.get('format', 'xlsx')

        suppliers = list(Supplier.objects.filter(
            organization=org, is_active=True
        ).values_list('name', flat=True))
        sales_cats = list(SalesCategory.objects.filter(
            organization=org, is_active=True
        ).values_list('name', flat=True))

        income_defaults = ['Cash', 'Card', 'Surcharge'] + sales_cats
        cogs_defaults = suppliers if suppliers else []

        # Generate date range: 2 years back from current month
        today = date.today()
        start_year = today.year - 2
        start_month = today.month + 1
        if start_month > 12:
            start_month = 1
            start_year += 1

        months = []
        for m_offset in range(24):
            y = start_year + (start_month + m_offset - 1) // 12
            mo = (start_month + m_offset - 1) % 12 + 1
            months.append((y, mo))

        if fmt == 'csv':
            return self._generate_csv(org, months, income_defaults, cogs_defaults)
        else:
            return self._generate_xlsx(org, months, income_defaults, cogs_defaults)

    def _generate_csv(self, org, months, income_defaults, cogs_defaults):
        """Generate CSV template in SkyT2 format."""
        output = io.StringIO()
        writer = csv.writer(output)

        for year, month in months:
            days = calendar.monthrange(year, month)[1]
            day_headers = ['Total'] + list(range(1, days + 1))
            # Pad to 31 columns
            while len(day_headers) < 32:
                day_headers.append('')

            writer.writerow(['Store Name', org.name])
            writer.writerow(['Month', f'{year}-{month:02d}'])
            writer.writerow([])

            # Income header
            writer.writerow(['Income'] + day_headers)
            for name in income_defaults:
                writer.writerow([name, ''] + [''] * 31)
            # Extra blank rows for new entries
            for _ in range(5):
                writer.writerow([''] + [''] * 32)
            writer.writerow(['Income Total'] + [''] * 32)
            writer.writerow([])

            # COGs header
            writer.writerow(['COGs (Supplier Costs)'] + day_headers)
            for name in cogs_defaults:
                writer.writerow([name, ''] + [''] * 31)
            writer.writerow(['Other', ''] + [''] * 31)
            for _ in range(5):
                writer.writerow([''] + [''] * 32)
            writer.writerow(['COGs Total'] + [''] * 32)
            writer.writerow([])
            writer.writerow([])

        content = output.getvalue()
        response = HttpResponse(content, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="OneOps_Import_Template_{org.name}.csv"'
        return response

    def _generate_xlsx(self, org, months, income_defaults, cogs_defaults):
        """Generate Excel template."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        section_font = Font(bold=True, size=12, color='FFFFFF')
        income_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cogs_fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
        total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_row_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        cogs_total_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
        grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        INCOME_ROWS = max(len(income_defaults) + 5, 15)
        COGS_ROWS = max(len(cogs_defaults) + 5, 15)

        # Reference sheet
        ref_ws = wb.create_sheet('Reference')
        ref_ws.cell(row=1, column=1, value='Income Channels').font = Font(bold=True, size=12)
        ref_ws.cell(row=1, column=3, value='COGs Suppliers').font = Font(bold=True, size=12)
        for i, name in enumerate(income_defaults, 2):
            ref_ws.cell(row=i, column=1, value=name)
        for i, name in enumerate(cogs_defaults, 2):
            ref_ws.cell(row=i, column=3, value=name)
        ref_ws.column_dimensions['A'].width = 20
        ref_ws.column_dimensions['C'].width = 20
        ref_ws.cell(row=1, column=5, value='How to Use').font = Font(bold=True, size=14)
        instructions = [
            '1. Fill in daily sales/costs in each monthly sheet',
            '2. Cash, Card rows = POS sales data',
            '3. Other income rows (Uber, DoorDash etc) = Other Sales',
            '4. COGs rows = Supplier costs',
            '5. New company names are auto-added to Store Settings',
            '6. 0 or blank = no data for that day (skipped)',
            '7. Existing data for same date will be overwritten',
            '8. You can also upload CSV files in the same format',
            '',
            'WARNING: Do not rename Cash/Card rows',
        ]
        for i, text in enumerate(instructions, 2):
            ref_ws.cell(row=i, column=5, value=text)
        ref_ws.column_dimensions['E'].width = 55

        for y, mo in months:
            sheet_name = f'{y}-{mo:02d}'
            ws = wb.create_sheet(sheet_name)
            days = calendar.monthrange(y, mo)[1]

            row = 1
            ws.cell(row=row, column=1, value='Income').font = section_font
            ws.cell(row=row, column=1).fill = income_fill
            ws.cell(row=row, column=2, value='Total').font = Font(bold=True, color='FFFFFF')
            ws.cell(row=row, column=2).fill = income_fill
            for d in range(1, days + 1):
                c = ws.cell(row=row, column=d + 2, value=d)
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = income_fill
                c.alignment = Alignment(horizontal='center')
            for d in range(days + 1, 32):
                ws.cell(row=row, column=d + 2).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

            for i in range(INCOME_ROWS):
                r = row + 1 + i
                nc = ws.cell(row=r, column=1)
                if i < len(income_defaults):
                    nc.value = income_defaults[i]
                nc.border = thin_border
                tc = ws.cell(row=r, column=2)
                tc.value = f'=SUM(C{r}:{get_column_letter(days + 2)}{r})'
                tc.font = Font(bold=True)
                tc.fill = total_fill
                tc.border = thin_border
                tc.number_format = '#,##0.00'
                for d in range(1, days + 1):
                    ws.cell(row=r, column=d + 2).border = thin_border
                    ws.cell(row=r, column=d + 2).number_format = '#,##0.00'
                for d in range(days + 1, 32):
                    ws.cell(row=r, column=d + 2).fill = grey_fill

            itr = row + 1 + INCOME_ROWS
            ws.cell(row=itr, column=1, value='Income Total').font = Font(bold=True)
            ws.cell(row=itr, column=1).fill = total_row_fill
            ws.cell(row=itr, column=1).border = thin_border
            for c in range(2, days + 3):
                cl = get_column_letter(c)
                cell = ws.cell(row=itr, column=c)
                cell.value = f'=SUM({cl}{row + 1}:{cl}{itr - 1})'
                cell.font = Font(bold=True)
                cell.fill = total_row_fill
                cell.border = thin_border
                cell.number_format = '#,##0.00'

            chr_ = itr + 2
            ws.cell(row=chr_, column=1, value='COGs').font = section_font
            ws.cell(row=chr_, column=1).fill = cogs_fill
            ws.cell(row=chr_, column=2, value='Total').font = Font(bold=True, color='FFFFFF')
            ws.cell(row=chr_, column=2).fill = cogs_fill
            for d in range(1, days + 1):
                c = ws.cell(row=chr_, column=d + 2, value=d)
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = cogs_fill
                c.alignment = Alignment(horizontal='center')
            for d in range(days + 1, 32):
                ws.cell(row=chr_, column=d + 2).fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')

            for i in range(COGS_ROWS):
                r = chr_ + 1 + i
                nc = ws.cell(row=r, column=1)
                if i < len(cogs_defaults):
                    nc.value = cogs_defaults[i]
                nc.border = thin_border
                tc = ws.cell(row=r, column=2)
                tc.value = f'=SUM(C{r}:{get_column_letter(days + 2)}{r})'
                tc.font = Font(bold=True)
                tc.fill = total_fill
                tc.border = thin_border
                tc.number_format = '#,##0.00'
                for d in range(1, days + 1):
                    ws.cell(row=r, column=d + 2).border = thin_border
                    ws.cell(row=r, column=d + 2).number_format = '#,##0.00'
                for d in range(days + 1, 32):
                    ws.cell(row=r, column=d + 2).fill = grey_fill

            ctr = chr_ + 1 + COGS_ROWS
            ws.cell(row=ctr, column=1, value='COGs Total').font = Font(bold=True)
            ws.cell(row=ctr, column=1).fill = cogs_total_fill
            ws.cell(row=ctr, column=1).border = thin_border
            for c in range(2, days + 3):
                cl = get_column_letter(c)
                cell = ws.cell(row=ctr, column=c)
                cell.value = f'=SUM({cl}{chr_ + 1}:{cl}{ctr - 1})'
                cell.font = Font(bold=True)
                cell.fill = cogs_total_fill
                cell.border = thin_border
                cell.number_format = '#,##0.00'

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 12
            for d in range(1, 32):
                ws.column_dimensions[get_column_letter(d + 2)].width = 9
            ws.freeze_panes = 'C2'

        wb.move_sheet('Reference', offset=-len(months))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="OneOps_Import_Template_{org.name}.xlsx"'
        return response
