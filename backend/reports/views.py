from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from django.db.models import Sum, Avg, Count, F, DecimalField
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime, date
from decimal import Decimal
from collections import defaultdict
import io

from .models import Report, GeneratedReport, SkyReport, StoreEvaluation, ProfitShare, PartnerShare
from .serializers import ReportSerializer, GeneratedReportSerializer, SkyReportSerializer, StoreEvaluationSerializer, ProfitShareSerializer
from closing.models import DailyClosing, ClosingHRCash, ClosingCashExpense, ClosingSupplierCost, SupplierMonthlyStatement, ClosingOtherSale, Holiday
from sales.models import Sales
from hr.models import Timesheet
from users.models import Organization
from users.permissions import IsManager, IsRegionalManager
from users.filters import OrganizationFilterBackend


class ReportViewSet(viewsets.ModelViewSet):
    """
    리포트 관리 ViewSet
    - list: 리포트 목록 조회
    - create: 리포트 생성
    - retrieve: 리포트 상세 조회
    - update/partial_update: 리포트 수정
    - destroy: 리포트 삭제
    - daily_store_report: 일일 매장 리포트
    - store_comparison: 매장 간 비교 리포트
    - sales_performance: 매출 성과 리포트
    """
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated, IsManager]
    filter_backends = [OrganizationFilterBackend]

    def get_queryset(self):
        """사용자의 조직에 해당하는 리포트만 조회"""
        queryset = super().get_queryset()
        return queryset.select_related('organization', 'created_by__user')

    def _get_organization(self, request):
        """사용자 프로필에서 조직 정보를 가져옵니다. CEO/HQ는 store_id로 전환 가능."""
        try:
            profile = request.user.profile
            # CEO/HQ can switch stores via store_id query param
            if profile.role in ['CEO', 'HQ']:
                store_id = request.query_params.get('store_id')
                if store_id:
                    from users.models import Organization
                    try:
                        return Organization.objects.get(id=store_id)
                    except Organization.DoesNotExist:
                        pass
            return profile.organization
        except Exception:
            return None

    @action(detail=False, methods=['get'])
    def daily_store_report(self, request):
        """
        일일 매장 리포트
        - 클로징 현황
        - 매출 현황
        - 차이 분석

        Query params:
        - date: YYYY-MM-DD (기본: 오늘)
        """
        try:
            date_str = request.query_params.get('date')
            target_date = timezone.localdate() if not date_str else date_str

            org = self._get_organization(request)

            if not org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 클로징 데이터
            closing = DailyClosing.objects.filter(
                organization=org,
                closing_date=target_date
            ).first()

            # 매출 데이터
            sales = Sales.objects.filter(
                organization=org,
                date=target_date
            ).aggregate(
                total=Sum('amount'),
                count=Count('id'),
                average=Avg('amount')
            )

            return Response({
                'date': target_date,
                'closing': {
                    'id': closing.id if closing else None,
                    'pos_total': float(closing.pos_total) if closing else 0,
                    'actual_total': float(closing.actual_total) if closing else 0,
                    'variance': float(closing.total_variance) if closing else 0,
                    'status': closing.status if closing else 'NOT_STARTED',
                    'hr_cash': float(sum(hc.amount for hc in closing.hr_cash_entries.all())) if closing else 0,
                    'bank_deposit': float(closing.bank_deposit) if closing else 0,
                },
                'sales': {
                    'total': float(sales['total'] or 0),
                    'transaction_count': sales['count'],
                    'average_transaction': float(sales['average'] or 0)
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated, IsRegionalManager])
    def store_comparison(self, request):
        """
        매장 간 비교 리포트
        - 선택한 날짜의 모든 매장 비교
        - 클로징 현황
        - 매출 현황

        Query params:
        - date: YYYY-MM-DD (기본: 오늘)
        """
        try:
            date_str = request.query_params.get('date')
            target_date = timezone.localdate() if not date_str else date_str

            # 현재 조직 및 하위 조직 찾기
            current_org = self._get_organization(request)
            if not current_org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 클로징 데이터 — pos_total / actual_total 은 property이므로 실제 필드를 합산
            closings = DailyClosing.objects.filter(
                closing_date=target_date
            ).values('organization__name', 'organization__id').annotate(
                pos_total=Sum(F('pos_card') + F('pos_cash'), output_field=DecimalField()),
                actual_total=Sum(F('actual_card') + F('actual_cash'), output_field=DecimalField()),
                variance=Sum(
                    F('actual_card') + F('actual_cash') - F('pos_card') - F('pos_cash'),
                    output_field=DecimalField()
                ),
            )

            # 매출 데이터
            sales_data = Sales.objects.filter(
                date=target_date
            ).values('organization__name', 'organization__id').annotate(
                total=Sum('amount'),
                count=Count('id'),
                average=Avg('amount')
            )

            # 데이터 결합
            stores = {}
            for closing in closings:
                org_id = closing['organization__id']
                stores[org_id] = {
                    'organization': closing['organization__name'],
                    'closing_pos_total': float(closing['pos_total'] or 0),
                    'closing_actual_total': float(closing['actual_total'] or 0),
                    'closing_variance': float(closing['variance'] or 0),
                    'sales_total': 0,
                    'sales_count': 0,
                    'sales_average': 0
                }

            for sale in sales_data:
                org_id = sale['organization__id']
                if org_id not in stores:
                    stores[org_id] = {
                        'organization': sale['organization__name'],
                        'closing_pos_total': 0,
                        'closing_actual_total': 0,
                        'closing_variance': 0,
                        'sales_total': 0,
                        'sales_count': 0,
                        'sales_average': 0
                    }
                stores[org_id]['sales_total'] = float(sale['total'] or 0)
                stores[org_id]['sales_count'] = sale['count']
                stores[org_id]['sales_average'] = float(sale['average'] or 0)

            return Response({
                'date': target_date,
                'stores': list(stores.values())
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def sales_performance(self, request):
        """
        매출 성과 리포트
        - 일일/주별/월별 성과
        - 목표 대비 달성도
        - 트렌드 분석

        Query params:
        - period: daily, weekly, monthly (기본: daily)
        - days: 분석 기간 (기본: 30)
        """
        try:
            period = request.query_params.get('period', 'daily')
            days = int(request.query_params.get('days', 30))

            end_date = timezone.localdate()
            start_date = end_date - timedelta(days=days)

            org = self._get_organization(request)

            if not org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 일별 매출 데이터
            daily_data = Sales.objects.filter(
                organization=org,
                date__range=[start_date, end_date]
            ).values('date').annotate(
                total=Sum('amount'),
                count=Count('id'),
                average=Avg('amount')
            ).order_by('date')

            # 기간별 집계
            if period == 'daily':
                performance_data = list(daily_data)
            elif period == 'weekly':
                weeks_dict = {}
                for item in daily_data:
                    week_start = item['date'] - timedelta(days=item['date'].weekday())
                    week_key = str(week_start)
                    if week_key not in weeks_dict:
                        weeks_dict[week_key] = {
                            'period': week_key,
                            'total': 0,
                            'count': 0,
                            'average': 0,
                            'day_count': 0
                        }
                    weeks_dict[week_key]['total'] += item['total']
                    weeks_dict[week_key]['count'] += item['count']
                    weeks_dict[week_key]['day_count'] += 1

                performance_data = [
                    {
                        'period': w['period'],
                        'total': float(w['total']),
                        'count': w['count'],
                        'average': float(w['total'] / w['day_count']) if w['day_count'] > 0 else 0
                    }
                    for w in sorted(weeks_dict.values(), key=lambda x: x['period'])
                ]
            else:  # monthly
                months_dict = {}
                for item in daily_data:
                    month_key = item['date'].strftime('%Y-%m')
                    if month_key not in months_dict:
                        months_dict[month_key] = {
                            'period': month_key,
                            'total': 0,
                            'count': 0,
                            'average': 0,
                            'day_count': 0
                        }
                    months_dict[month_key]['total'] += item['total']
                    months_dict[month_key]['count'] += item['count']
                    months_dict[month_key]['day_count'] += 1

                performance_data = [
                    {
                        'period': m['period'],
                        'total': float(m['total']),
                        'count': m['count'],
                        'average': float(m['total'] / m['day_count']) if m['day_count'] > 0 else 0
                    }
                    for m in sorted(months_dict.values(), key=lambda x: x['period'])
                ]

            # 통계 계산
            total_sales = sum(p['total'] for p in performance_data)
            avg_sales = total_sales / len(performance_data) if performance_data else 0
            max_sales = max((p['total'] for p in performance_data), default=0)
            min_sales = min((p['total'] for p in performance_data), default=0)

            return Response({
                'period': period,
                'start_date': start_date,
                'end_date': end_date,
                'statistics': {
                    'total': float(total_sales),
                    'average': float(avg_sales),
                    'max': float(max_sales),
                    'min': float(min_sales),
                    'trend': 'up' if len(performance_data) > 1 and performance_data[-1]['total'] > performance_data[0]['total'] else 'down'
                },
                'performance_data': performance_data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def cash_report(self, request):
        """
        현금 흐름 리포트
        - 현금 매출, 은행 입금, HR 현금, 현금 지출, 잔액

        Query params:
        - date: YYYY-MM-DD (단일 날짜)
        - start_date / end_date: YYYY-MM-DD (기간 조회)
        """
        try:
            org = self._get_organization(request)

            if not org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 날짜 파라미터 처리
            single_date = request.query_params.get('date')
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            if single_date:
                start = datetime.strptime(single_date, '%Y-%m-%d').date()
                end = start
            elif start_date_str and end_date_str:
                start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                start = timezone.localdate()
                end = start

            # 기간 내 클로징 데이터 조회
            closings = DailyClosing.objects.filter(
                organization=org,
                closing_date__range=[start, end]
            ).prefetch_related('hr_cash_entries', 'cash_expenses').order_by('closing_date')

            # 클로징을 날짜별 딕셔너리로
            closing_map = {c.closing_date: c for c in closings}

            # 날짜 리스트 생성
            daily_reports = []
            totals = {
                'cash_sales': Decimal('0'),
                'bank_deposit': Decimal('0'),
                'hr_cash_total': Decimal('0'),
                'cash_expenses_total': Decimal('0'),
                'balance': Decimal('0'),
            }

            today = timezone.localdate()
            # 미래 날짜는 오늘까지만
            effective_end = min(end, today)

            current = start
            while current <= effective_end:
                closing = closing_map.get(current)

                if not closing:
                    daily_reports.append({
                        'date': str(current),
                        'cash_sales': 0,
                        'bank_deposit': 0,
                        'hr_cash_total': 0,
                        'hr_cash_entries': [],
                        'cash_expenses_total': 0,
                        'cash_expenses': [],
                        'balance': 0,
                        'has_data': False,
                        'status': None,
                    })
                    current += timedelta(days=1)
                    continue

                # HR Cash 항목
                hr_entries = []
                hr_total = Decimal('0')
                for entry in closing.hr_cash_entries.all():
                    hr_total += entry.amount
                    hr_entries.append({
                        'id': entry.id,
                        'recipient_name': entry.recipient_name or '',
                        'amount': float(entry.amount),
                        'notes': entry.notes or '',
                        'photo': request.build_absolute_uri(entry.photo.url) if entry.photo else None,
                    })

                # Implicit HR Cash: if no entries, treat remaining cash as HR cash
                if not hr_entries:
                    implicit_hr = closing.actual_cash - closing.bank_deposit
                    if implicit_hr > 0:
                        hr_total = implicit_hr
                        hr_entries.append({
                            'id': None,
                            'recipient_name': 'Unallocated',
                            'amount': float(implicit_hr),
                            'notes': 'Auto-calculated (actual cash − bank deposit)',
                            'photo': None,
                        })

                # Cash Expense 항목
                expense_entries = []
                expenses_total = Decimal('0')
                for expense in closing.cash_expenses.all():
                    expenses_total += expense.amount
                    expense_entries.append({
                        'id': expense.id,
                        'category': expense.category,
                        'category_display': expense.get_category_display(),
                        'reason': expense.reason,
                        'amount': float(expense.amount),
                        'attachment': request.build_absolute_uri(expense.attachment.url) if expense.attachment else None,
                    })

                balance = closing.actual_cash - closing.bank_deposit - hr_total - expenses_total

                day_data = {
                    'date': str(current),
                    'cash_sales': float(closing.actual_cash),
                    'bank_deposit': float(closing.bank_deposit),
                    'hr_cash_total': float(hr_total),
                    'hr_cash_entries': hr_entries,
                    'cash_expenses_total': float(expenses_total),
                    'cash_expenses': expense_entries,
                    'balance': float(balance),
                    'has_data': True,
                    'status': closing.status,
                }
                daily_reports.append(day_data)

                totals['cash_sales'] += closing.actual_cash
                totals['bank_deposit'] += closing.bank_deposit
                totals['hr_cash_total'] += hr_total
                totals['cash_expenses_total'] += expenses_total
                totals['balance'] += balance

                current += timedelta(days=1)

            # 미완료 일수 (오늘까지 중 has_data=False인 날)
            missing_count = sum(1 for d in daily_reports if not d['has_data'])

            # 전달 HR Balance 계산 (조회 시작일 이전 모든 데이터)
            # Start with store's initial cash balance
            previous_hr_balance = Decimal(str(org.initial_cash_balance or 0))

            prev_closings = DailyClosing.objects.filter(
                organization=org,
                closing_date__lt=start,
            ).prefetch_related('hr_cash_entries', 'cash_expenses')

            for pc in prev_closings:
                pc_hr = sum(e.amount for e in pc.hr_cash_entries.all())
                if not pc.hr_cash_entries.exists():
                    implicit = pc.actual_cash - pc.bank_deposit
                    if implicit > 0:
                        pc_hr = implicit
                pc_exp = sum(e.amount for e in pc.cash_expenses.all())
                previous_hr_balance += pc_hr - pc_exp

            return Response({
                'period': {
                    'start_date': str(start),
                    'end_date': str(end),
                    'days': (end - start).days + 1,
                },
                'totals': {k: float(v) for k, v in totals.items()},
                'previous_hr_balance': float(previous_hr_balance),
                'missing_count': missing_count,
                'daily_reports': daily_reports,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def hr_cash_report(self, request):
        """
        HR Cash 추적 리포트
        - 수령인별 집계, 카테고리별 분류, 일별 상세

        Query params:
        - date: YYYY-MM-DD (단일 날짜)
        - start_date / end_date: YYYY-MM-DD (기간 조회)
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 날짜 파라미터 처리
            single_date = request.query_params.get('date')
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            if single_date:
                start = datetime.strptime(single_date, '%Y-%m-%d').date()
                end = start
            elif start_date_str and end_date_str:
                start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                start = timezone.localdate()
                end = start

            # 클로징 데이터 조회
            closings = DailyClosing.objects.filter(
                organization=org,
                closing_date__range=[start, end]
            ).prefetch_related('hr_cash_entries', 'cash_expenses').order_by('closing_date')

            # 집계용
            hr_cash_total = Decimal('0')
            expenses_total = Decimal('0')
            entry_count = 0
            days_with_data = 0
            recipient_agg = defaultdict(lambda: {'total_amount': Decimal('0'), 'count': 0})
            category_agg = defaultdict(lambda: {'total_amount': Decimal('0'), 'count': 0, 'category_display': ''})
            daily_reports = []

            for closing in closings:
                hr_total_day = Decimal('0')
                hr_entries = []
                for entry in closing.hr_cash_entries.all():
                    hr_total_day += entry.amount
                    entry_count += 1
                    name = entry.recipient_name or 'Unknown'
                    recipient_agg[name]['total_amount'] += entry.amount
                    recipient_agg[name]['count'] += 1
                    hr_entries.append({
                        'id': entry.id,
                        'recipient_name': entry.recipient_name or '',
                        'amount': float(entry.amount),
                        'notes': entry.notes or '',
                        'photo': request.build_absolute_uri(entry.photo.url) if entry.photo else None,
                    })

                # Implicit HR Cash: if no entries, treat remaining cash as HR cash
                if not hr_entries:
                    implicit_hr = closing.actual_cash - closing.bank_deposit
                    if implicit_hr > 0:
                        hr_total_day = implicit_hr
                        entry_count += 1
                        recipient_agg['Unallocated']['total_amount'] += implicit_hr
                        recipient_agg['Unallocated']['count'] += 1
                        hr_entries.append({
                            'id': None,
                            'recipient_name': 'Unallocated',
                            'amount': float(implicit_hr),
                            'notes': 'Auto-calculated',
                            'photo': None,
                        })

                exp_total_day = Decimal('0')
                expense_entries = []
                for expense in closing.cash_expenses.all():
                    exp_total_day += expense.amount
                    entry_count += 1
                    cat = expense.category
                    category_agg[cat]['total_amount'] += expense.amount
                    category_agg[cat]['count'] += 1
                    category_agg[cat]['category_display'] = expense.get_category_display()
                    expense_entries.append({
                        'id': expense.id,
                        'category': expense.category,
                        'category_display': expense.get_category_display(),
                        'reason': expense.reason,
                        'amount': float(expense.amount),
                        'attachment': request.build_absolute_uri(expense.attachment.url) if expense.attachment else None,
                    })

                if hr_entries or expense_entries:
                    days_with_data += 1
                    hr_cash_total += hr_total_day
                    expenses_total += exp_total_day
                    daily_reports.append({
                        'date': str(closing.closing_date),
                        'hr_cash_total': float(hr_total_day),
                        'hr_cash_entries': hr_entries,
                        'cash_expenses_total': float(exp_total_day),
                        'cash_expenses': expense_entries,
                    })

            grand_total = hr_cash_total + expenses_total
            average_daily = float(grand_total / days_with_data) if days_with_data > 0 else 0

            # 수령인 요약
            recipient_summary = sorted(
                [{'recipient_name': name, 'total_amount': float(v['total_amount']), 'count': v['count']}
                 for name, v in recipient_agg.items()],
                key=lambda x: x['total_amount'], reverse=True
            )

            # 카테고리 요약
            category_summary = sorted(
                [{'category': cat, 'category_display': v['category_display'],
                  'total_amount': float(v['total_amount']), 'count': v['count']}
                 for cat, v in category_agg.items()],
                key=lambda x: x['total_amount'], reverse=True
            )

            return Response({
                'period': {
                    'start_date': str(start),
                    'end_date': str(end),
                    'days': (end - start).days + 1,
                },
                'totals': {
                    'hr_cash_total': float(hr_cash_total),
                    'expenses_total': float(expenses_total),
                    'grand_total': float(grand_total),
                    'entry_count': entry_count,
                    'average_daily': average_daily,
                },
                'recipient_summary': recipient_summary,
                'category_summary': category_summary,
                'daily_reports': daily_reports,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def supply_report(self, request):
        """
        Supply cost report
        - Monthly: supplier costs aggregated by supplier for settlement
        - Daily: day-by-day breakdown with per-supplier detail

        Query params:
        - period: monthly (default), daily
        - month: YYYY-MM (for monthly, default: current month)
        - start_date / end_date: YYYY-MM-DD (for daily)
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)

            period = request.query_params.get('period', 'monthly')

            if period == 'monthly':
                month_str = request.query_params.get('month')
                if month_str:
                    parts = month_str.split('-')
                    year, month = int(parts[0]), int(parts[1])
                else:
                    today = timezone.localdate()
                    year, month = today.year, today.month

                from datetime import date as dt_date
                start = dt_date(year, month, 1)
                if month == 12:
                    end = dt_date(year + 1, 1, 1)
                else:
                    end = dt_date(year, month + 1, 1)

                # Aggregate by supplier
                supplier_data = ClosingSupplierCost.objects.filter(
                    closing__organization=org,
                    closing__closing_date__gte=start,
                    closing__closing_date__lt=end,
                ).values(
                    'supplier__id', 'supplier__name', 'supplier__code'
                ).annotate(
                    total=Sum('amount'),
                    entry_count=Count('id'),
                ).order_by('supplier__name')

                # Get statements for reconciliation
                statements = SupplierMonthlyStatement.objects.filter(
                    organization=org, year=year, month=month
                )
                statement_map = {s.supplier_id: s for s in statements}

                suppliers_list = []
                grand_total = Decimal('0')
                for item in supplier_data:
                    supplier_id = item['supplier__id']
                    stmt = statement_map.get(supplier_id)
                    suppliers_list.append({
                        'supplier_id': supplier_id,
                        'supplier_name': item['supplier__name'] or 'Unknown',
                        'supplier_code': item['supplier__code'] or '',
                        'total': float(item['total'] or 0),
                        'entry_count': item['entry_count'],
                        'statement': {
                            'id': stmt.id,
                            'statement_total': float(stmt.statement_total),
                            'our_total': float(stmt.our_total),
                            'status': stmt.status,
                            'variance': float(stmt.variance),
                        } if stmt else None,
                    })
                    grand_total += item['total'] or 0

                return Response({
                    'period': 'monthly',
                    'month': f'{year}-{month:02d}',
                    'grand_total': float(grand_total),
                    'supplier_count': len(suppliers_list),
                    'suppliers': suppliers_list,
                }, status=status.HTTP_200_OK)

            else:  # daily
                start_str = request.query_params.get('start_date')
                end_str = request.query_params.get('end_date')
                if start_str and end_str:
                    start = datetime.strptime(start_str, '%Y-%m-%d').date()
                    end = datetime.strptime(end_str, '%Y-%m-%d').date()
                else:
                    end = timezone.localdate()
                    start = end - timedelta(days=30)

                entries = ClosingSupplierCost.objects.filter(
                    closing__organization=org,
                    closing__closing_date__range=[start, end],
                ).select_related('supplier', 'closing').order_by('closing__closing_date', 'supplier__name')

                # Group by date
                from collections import defaultdict
                daily_map = defaultdict(lambda: defaultdict(list))
                for entry in entries:
                    d = str(entry.closing.closing_date)
                    s_name = entry.supplier.name if entry.supplier else 'Unknown'
                    s_id = entry.supplier_id
                    daily_map[d][(s_id, s_name)].append({
                        'id': entry.id,
                        'amount': float(entry.amount),
                        'description': entry.description or '',
                        'invoice_number': entry.invoice_number or '',
                    })

                daily_reports = []
                grand_total = Decimal('0')
                for d in sorted(daily_map.keys()):
                    day_total = 0
                    suppliers = []
                    for (s_id, s_name), items in daily_map[d].items():
                        subtotal = sum(i['amount'] for i in items)
                        suppliers.append({
                            'supplier_id': s_id,
                            'supplier_name': s_name,
                            'entries': items,
                            'subtotal': subtotal,
                        })
                        day_total += subtotal
                    daily_reports.append({
                        'date': d,
                        'total': day_total,
                        'suppliers': suppliers,
                    })
                    grand_total += Decimal(str(day_total))

                return Response({
                    'period': 'daily',
                    'start_date': str(start),
                    'end_date': str(end),
                    'grand_total': float(grand_total),
                    'daily_reports': daily_reports,
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def supply_detail(self, request):
        """
        Drill-down: individual entries for a specific supplier in a month.
        Used when investigating mismatches.

        Query params:
        - supplier_id: required
        - month: YYYY-MM (default: current month)
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)

            supplier_id = request.query_params.get('supplier_id')
            if not supplier_id:
                return Response({'error': 'supplier_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

            month_str = request.query_params.get('month')
            if month_str:
                parts = month_str.split('-')
                year, month = int(parts[0]), int(parts[1])
            else:
                today = timezone.localdate()
                year, month = today.year, today.month

            from datetime import date as dt_date
            start = dt_date(year, month, 1)
            if month == 12:
                end = dt_date(year + 1, 1, 1)
            else:
                end = dt_date(year, month + 1, 1)

            entries = ClosingSupplierCost.objects.filter(
                closing__organization=org,
                closing__closing_date__gte=start,
                closing__closing_date__lt=end,
                supplier_id=supplier_id,
            ).select_related('supplier', 'closing').order_by('closing__closing_date')

            # Group by date
            from collections import defaultdict
            by_date = defaultdict(list)
            total = Decimal('0')
            for entry in entries:
                d = str(entry.closing.closing_date)
                by_date[d].append({
                    'id': entry.id,
                    'amount': float(entry.amount),
                    'description': entry.description or '',
                    'invoice_number': entry.invoice_number or '',
                })
                total += entry.amount or 0

            daily_entries = []
            for d in sorted(by_date.keys()):
                items = by_date[d]
                daily_entries.append({
                    'date': d,
                    'subtotal': sum(i['amount'] for i in items),
                    'entries': items,
                })

            supplier = entries.first().supplier if entries.exists() else None

            return Response({
                'supplier_id': int(supplier_id),
                'supplier_name': supplier.name if supplier else 'Unknown',
                'month': f'{year}-{month:02d}',
                'total': float(total),
                'entry_count': entries.count(),
                'daily_entries': daily_entries,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def sales_report(self, request):
        """
        Sales report based on DailyClosing data
        - Daily/Weekly/Monthly breakdown with card/cash split

        Query params (date-based — preferred):
        - date: single date (YYYY-MM-DD)
        - start_date & end_date: date range
        Legacy params:
        - period: daily (default), weekly, monthly
        - days: number of days back (default: 30)
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)

            period = request.query_params.get('period', 'daily')

            # Date-based params (new)
            date_param = request.query_params.get('date')
            start_param = request.query_params.get('start_date')
            end_param = request.query_params.get('end_date')

            today = timezone.localdate()

            if start_param and end_param:
                start_date = datetime.strptime(start_param, '%Y-%m-%d').date()
                end_date = min(datetime.strptime(end_param, '%Y-%m-%d').date(), today)
            elif date_param:
                start_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                end_date = start_date
            else:
                # Legacy fallback
                days = int(request.query_params.get('days', 30))
                end_date = today
                start_date = end_date - timedelta(days=days)

            # Use _compute_store_sales which includes Other Sales
            result = self._compute_store_sales(org, start_date, end_date)

            # Build daily data with status/variance (not in _compute_store_sales)
            closings = DailyClosing.objects.filter(
                organization=org,
                closing_date__range=[start_date, end_date]
            ).order_by('closing_date')

            from closing.models import ClosingOtherSale
            daily_data = []
            for c in closings:
                other_sales_total = ClosingOtherSale.objects.filter(closing=c).aggregate(
                    total=Sum('amount'))['total'] or Decimal('0')
                total = float(c.actual_card + c.actual_cash) + float(other_sales_total)
                daily_data.append({
                    'date': str(c.closing_date),
                    'pos_total': float(c.pos_total),
                    'actual_total': round(total, 2),
                    'card_sales': float(c.actual_card),
                    'cash_sales': float(c.actual_cash),
                    'other_sales': float(other_sales_total),
                    'variance': float(c.total_variance),
                    'status': c.status,
                })

            # Aggregate by period
            if period == 'weekly':
                weeks_dict = {}
                for item in daily_data:
                    d = datetime.strptime(item['date'], '%Y-%m-%d').date()
                    week_start = d - timedelta(days=(d.weekday() + 1) % 7)
                    week_key = str(week_start)
                    if week_key not in weeks_dict:
                        weeks_dict[week_key] = {
                            'period': week_key, 'actual_total': 0, 'card_sales': 0,
                            'cash_sales': 0, 'other_sales': 0, 'variance': 0, 'day_count': 0,
                        }
                    weeks_dict[week_key]['actual_total'] += item['actual_total']
                    weeks_dict[week_key]['card_sales'] += item['card_sales']
                    weeks_dict[week_key]['cash_sales'] += item['cash_sales']
                    weeks_dict[week_key]['other_sales'] += item['other_sales']
                    weeks_dict[week_key]['variance'] += item['variance']
                    weeks_dict[week_key]['day_count'] += 1
                result_data = sorted(weeks_dict.values(), key=lambda x: x['period'])
            elif period == 'monthly':
                months_dict = {}
                for item in daily_data:
                    month_key = item['date'][:7]
                    if month_key not in months_dict:
                        months_dict[month_key] = {
                            'period': month_key, 'actual_total': 0, 'card_sales': 0,
                            'cash_sales': 0, 'other_sales': 0, 'variance': 0, 'day_count': 0,
                        }
                    months_dict[month_key]['actual_total'] += item['actual_total']
                    months_dict[month_key]['card_sales'] += item['card_sales']
                    months_dict[month_key]['cash_sales'] += item['cash_sales']
                    months_dict[month_key]['other_sales'] += item['other_sales']
                    months_dict[month_key]['variance'] += item['variance']
                    months_dict[month_key]['day_count'] += 1
                result_data = sorted(months_dict.values(), key=lambda x: x['period'])
            else:
                result_data = daily_data

            # Statistics from _compute_store_sales (includes other_total, card_total, cash_total)
            return Response({
                'period': period,
                'start_date': str(start_date),
                'end_date': str(end_date),
                'statistics': {
                    'total_sales': result['total_sales'],
                    'average_daily': result['average_daily'],
                    'card_total': result['card_total'],
                    'cash_total': result['cash_total'],
                    'other_total': result['other_total'],
                    'highest': result['highest'],
                    'lowest': result['lowest'],
                    'trend': result['trend'],
                    'trend_percentage': result['trend_percentage'],
                },
                'data': result_data,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _get_accessible_stores(self, request):
        """Return stores the user can access based on role."""
        profile = request.user.profile
        if profile.role in ['CEO', 'HQ', 'ADMIN']:
            return Organization.objects.filter(level='STORE').order_by('name')
        elif profile.role in ['REGIONAL_MANAGER', 'SENIOR_MANAGER']:
            return Organization.objects.filter(
                level='STORE', parent=profile.organization
            ).order_by('name') | Organization.objects.filter(
                id=profile.organization_id, level='STORE'
            ).order_by('name')
        else:
            return Organization.objects.filter(id=profile.organization_id)

    def _compute_store_sales(self, org, start_date, end_date):
        """Compute sales data for a single store. Reused by sales_report and multi_store_sales."""
        from closing.models import ClosingOtherSale

        closings = DailyClosing.objects.filter(
            organization=org,
            closing_date__range=[start_date, end_date]
        ).order_by('closing_date')

        daily_data = []
        for c in closings:
            other_sales_total = ClosingOtherSale.objects.filter(closing=c).aggregate(
                total=Sum('amount'))['total'] or 0
            total = float(c.pos_card + c.pos_cash) + float(other_sales_total)
            daily_data.append({
                'date': str(c.closing_date),
                'actual_total': total,
                'card_sales': float(c.actual_card),
                'cash_sales': float(c.actual_cash),
                'other_sales': float(other_sales_total),
                'variance': float(c.total_variance),
                'status': c.status,
            })

        totals = [d['actual_total'] for d in daily_data]
        total_sales = sum(totals)
        num_days = max(len(daily_data), 1)
        avg_daily = total_sales / num_days

        card_total = sum(d['card_sales'] for d in daily_data)
        cash_total = sum(d['cash_sales'] for d in daily_data)
        other_total = sum(d['other_sales'] for d in daily_data)

        best = max(daily_data, key=lambda d: d['actual_total'], default=None)
        worst = min(daily_data, key=lambda d: d['actual_total'], default=None)

        # Trend
        if len(daily_data) >= 2:
            mid = len(daily_data) // 2
            first_half = sum(d['actual_total'] for d in daily_data[:mid])
            second_half = sum(d['actual_total'] for d in daily_data[mid:])
            trend_pct = round(((second_half - first_half) / max(first_half, 1)) * 100, 1)
            trend = 'up' if trend_pct > 0 else 'down' if trend_pct < 0 else 'stable'
        else:
            trend, trend_pct = 'stable', 0

        # Day-of-week averages
        dow_map = defaultdict(list)
        for d in daily_data:
            dt = datetime.strptime(d['date'], '%Y-%m-%d').date()
            dow_map[dt.strftime('%a')].append(d['actual_total'])

        day_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        day_of_week_avg = []
        for day in day_order:
            vals = dow_map.get(day, [])
            day_of_week_avg.append({
                'day': day,
                'avg_sales': round(sum(vals) / max(len(vals), 1), 2) if vals else 0,
                'count': len(vals),
            })

        return {
            'total_sales': round(total_sales, 2),
            'average_daily': round(avg_daily, 2),
            'card_total': round(card_total, 2),
            'cash_total': round(cash_total, 2),
            'other_total': round(other_total, 2),
            'highest': {'date': best['date'] if best else None, 'amount': best['actual_total'] if best else 0},
            'lowest': {'date': worst['date'] if worst else None, 'amount': worst['actual_total'] if worst else 0},
            'trend': trend,
            'trend_percentage': trend_pct,
            'data': daily_data,
            'day_of_week_avg': day_of_week_avg,
        }

    @action(detail=False, methods=['get'])
    def multi_store_sales(self, request):
        """
        Multi-store sales comparison.
        Query params:
        - start_date, end_date: YYYY-MM-DD (required)
        - store_ids: comma-separated store IDs (optional, all if omitted)
        """
        try:
            start_str = request.query_params.get('start_date')
            end_str = request.query_params.get('end_date')
            if not start_str or not end_str:
                return Response({'error': 'start_date and end_date required.'}, status=status.HTTP_400_BAD_REQUEST)

            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

            stores = self._get_accessible_stores(request)
            store_ids_param = request.query_params.get('store_ids')
            if store_ids_param:
                ids = [int(x) for x in store_ids_param.split(',') if x.strip()]
                stores = stores.filter(id__in=ids)

            store_results = []
            for store in stores:
                result = self._compute_store_sales(store, start_date, end_date)
                store_results.append({
                    'store_id': store.id,
                    'store_name': store.name,
                    **result,
                })

            # Combined totals
            combined = self._compute_store_sales_combined(store_results)

            return Response({
                'start_date': str(start_date),
                'end_date': str(end_date),
                'stores': store_results,
                'combined': combined,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _compute_store_sales_combined(self, store_results):
        """Aggregate multiple store results into combined totals."""
        if not store_results:
            return {'total_sales': 0, 'average_daily': 0, 'card_total': 0, 'cash_total': 0}

        total = sum(s['total_sales'] for s in store_results)
        card = sum(s['card_total'] for s in store_results)
        cash = sum(s['cash_total'] for s in store_results)

        # Combine day-of-week averages
        dow_combined = defaultdict(list)
        for s in store_results:
            for d in s.get('day_of_week_avg', []):
                if d['avg_sales'] > 0:
                    dow_combined[d['day']].append(d['avg_sales'])

        day_order = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        day_of_week_avg = []
        for day in day_order:
            vals = dow_combined.get(day, [])
            day_of_week_avg.append({
                'day': day,
                'avg_sales': round(sum(vals), 2) if vals else 0,
            })

        return {
            'total_sales': round(total, 2),
            'average_daily': round(total / max(len(store_results), 1), 2),
            'card_total': round(card, 2),
            'cash_total': round(cash, 2),
            'day_of_week_avg': day_of_week_avg,
        }

    @action(detail=False, methods=['get'])
    def ai_insights(self, request):
        """
        AI-powered sales analysis using Claude API.
        Query params:
        - start_date, end_date: YYYY-MM-DD (required)
        - store_id: specific store (optional)
        """
        try:
            start_str = request.query_params.get('start_date')
            end_str = request.query_params.get('end_date')
            if not start_str or not end_str:
                return Response({'error': 'start_date and end_date required.'}, status=status.HTTP_400_BAD_REQUEST)

            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

            store_id = request.query_params.get('store_id')
            if store_id:
                try:
                    org = Organization.objects.get(id=store_id)
                except Organization.DoesNotExist:
                    return Response({'error': 'Store not found.'}, status=status.HTTP_404_NOT_FOUND)
                store_name = org.name
            else:
                org = self._get_organization(request)
                store_name = org.name if org else 'Unknown'

            if not org:
                return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)

            result = self._compute_store_sales(org, start_date, end_date)

            from .services import generate_sales_insights
            insights = generate_sales_insights(result, store_name, str(start_date), str(end_date))

            if insights is None:
                return Response({
                    'insights': None,
                    'error': 'AI analysis unavailable. ANTHROPIC_API_KEY not configured.',
                }, status=status.HTTP_200_OK)

            return Response({
                'insights': insights,
                'generated_at': timezone.now().isoformat(),
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def chart_data(self, request):
        """
        차트 데이터 (일별 Sales/QTY/Labour/COGS + 작년 비교)

        Query params:
        - start_date: YYYY-MM-DD
        - end_date: YYYY-MM-DD
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response({'error': 'Organization not found.'}, status=status.HTTP_400_BAD_REQUEST)

            start_str = request.query_params.get('start_date')
            end_str = request.query_params.get('end_date')

            if not start_str or not end_str:
                return Response({'error': 'start_date and end_date required.'}, status=status.HTTP_400_BAD_REQUEST)

            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

            # Last year same period (same month/day, 1 year back)
            try:
                ly_start = start_date.replace(year=start_date.year - 1)
            except ValueError:  # Feb 29 → Feb 28
                ly_start = start_date.replace(year=start_date.year - 1, day=28)
            try:
                ly_end = end_date.replace(year=end_date.year - 1)
            except ValueError:
                ly_end = end_date.replace(year=end_date.year - 1, day=28)

            # Sales aggregation by date (try Sales model first, fallback to DailyClosing POS data)
            sales_agg = {
                row['date']: {'total': float(row['total'] or 0), 'qty': row['qty']}
                for row in Sales.objects.filter(
                    organization=org, date__range=[start_date, end_date]
                ).values('date').annotate(total=Sum('amount'), qty=Count('id'))
            }

            # Fallback: if no Sales data, use DailyClosing POS totals + Other Sales
            if not sales_agg:
                from closing.models import ClosingOtherSale
                other_sales_agg = {
                    row['closing__closing_date']: float(row['total'] or 0)
                    for row in ClosingOtherSale.objects.filter(
                        closing__organization=org,
                        closing__closing_date__range=[start_date, end_date]
                    ).values('closing__closing_date').annotate(total=Sum('amount'))
                }
                for dc in DailyClosing.objects.filter(
                    organization=org, closing_date__range=[start_date, end_date]
                ):
                    pos_total = float(dc.pos_card or 0) + float(dc.pos_cash or 0)
                    other_total = other_sales_agg.get(dc.closing_date, 0)
                    total = pos_total + other_total
                    if total > 0:
                        sales_agg[dc.closing_date] = {'total': total, 'qty': 0}

            ly_sales_agg = {
                row['date']: {'total': float(row['total'] or 0), 'qty': row['qty']}
                for row in Sales.objects.filter(
                    organization=org, date__range=[ly_start, ly_end]
                ).values('date').annotate(total=Sum('amount'), qty=Count('id'))
            }

            # Fallback for last year too
            if not ly_sales_agg:
                from closing.models import ClosingOtherSale as LyOtherSale
                ly_other_sales_agg = {
                    row['closing__closing_date']: float(row['total'] or 0)
                    for row in LyOtherSale.objects.filter(
                        closing__organization=org,
                        closing__closing_date__range=[ly_start, ly_end]
                    ).values('closing__closing_date').annotate(total=Sum('amount'))
                }
                for dc in DailyClosing.objects.filter(
                    organization=org, closing_date__range=[ly_start, ly_end]
                ):
                    pos_total = float(dc.pos_card or 0) + float(dc.pos_cash or 0)
                    other_total = ly_other_sales_agg.get(dc.closing_date, 0)
                    total = pos_total + other_total
                    if total > 0:
                        ly_sales_agg[dc.closing_date] = {'total': total, 'qty': 0}

            # COGS aggregation by date
            cogs_agg = {
                row['closing__closing_date']: float(row['total'] or 0)
                for row in ClosingSupplierCost.objects.filter(
                    closing__organization=org,
                    closing__closing_date__range=[start_date, end_date]
                ).values('closing__closing_date').annotate(total=Sum('amount'))
            }

            ly_cogs_agg = {
                row['closing__closing_date']: float(row['total'] or 0)
                for row in ClosingSupplierCost.objects.filter(
                    closing__organization=org,
                    closing__closing_date__range=[ly_start, ly_end]
                ).values('closing__closing_date').annotate(total=Sum('amount'))
            }

            # Labour hours: worked_hours is a property, group in Python
            timesheets = Timesheet.objects.filter(
                user__organization=org,
                date__range=[start_date, end_date],
                check_in__isnull=False, check_out__isnull=False,
            )
            labour_agg = defaultdict(float)
            for ts in timesheets:
                labour_agg[ts.date] += ts.worked_hours

            ly_timesheets = Timesheet.objects.filter(
                user__organization=org,
                date__range=[ly_start, ly_end],
                check_in__isnull=False, check_out__isnull=False,
            )
            ly_labour_agg = defaultdict(float)
            for ts in ly_timesheets:
                ly_labour_agg[ts.date] += ts.worked_hours

            # Build response: one entry per day
            days = []
            current = start_date
            day_idx = 0
            while current <= end_date:
                ly_date = ly_start + timedelta(days=day_idx)
                if ly_date > ly_end:
                    ly_date = ly_end  # clamp to avoid overflow into next month
                s = sales_agg.get(current, {'total': 0, 'qty': 0})
                ly_s = ly_sales_agg.get(ly_date, {'total': 0, 'qty': 0})

                days.append({
                    'date': str(current),
                    'label': current.strftime('%a %d/%m'),
                    'sales': round(s['total'], 2),
                    'qty': s['qty'],
                    'labour_hours': round(labour_agg.get(current, 0), 1),
                    'cogs': round(cogs_agg.get(current, 0), 2),
                    'ly_sales': round(ly_s['total'], 2),
                    'ly_qty': ly_s['qty'],
                    'ly_labour_hours': round(ly_labour_agg.get(ly_date, 0), 1),
                    'ly_cogs': round(ly_cogs_agg.get(ly_date, 0), 2),
                })
                current += timedelta(days=1)
                day_idx += 1

            return Response({'days': days}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def holiday_report(self, request):
        """
        Holiday report — sales + labor analysis for all holidays in a given year.

        Query params:
        - year: target year (default: current year)
        - store_id: optional store override for CEO/HQ
        """
        try:
            org = self._get_organization(request)
            if not org:
                return Response(
                    {'error': '조직 정보를 찾을 수 없습니다.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            current_year = timezone.now().year
            year = int(request.query_params.get('year', current_year))

            holidays = Holiday.objects.filter(year=year).order_by('start_date')
            available_years = sorted(set(
                Holiday.objects.values_list('year', flat=True)
            ))

            holiday_results = []
            total_holiday_sales = 0
            total_holiday_days = 0
            total_normal_sales_for_avg = 0
            total_normal_days = 0
            total_holiday_hours = 0

            for h in holidays:
                h_start = h.start_date
                h_end = h.end_date
                duration = (h_end - h_start).days + 1

                # --- Sales ---
                closings = DailyClosing.objects.filter(
                    organization=org,
                    closing_date__range=[h_start, h_end],
                )
                sales_agg = closings.aggregate(
                    card=Sum('actual_card'),
                    cash=Sum('actual_cash'),
                )
                other_agg = ClosingOtherSale.objects.filter(
                    closing__organization=org,
                    closing__closing_date__range=[h_start, h_end],
                ).aggregate(total=Sum('amount'))

                card = float(sales_agg['card'] or 0)
                cash = float(sales_agg['cash'] or 0)
                other_total = float(other_agg['total'] or 0)
                holiday_total = card + cash + other_total
                days_with_data = closings.count()
                avg_daily = holiday_total / max(days_with_data, 1)

                # --- Normal period comparison (2 weeks before holiday) ---
                compare_start = h_start - timedelta(days=14)
                compare_end = h_start - timedelta(days=1)
                compare_closings = DailyClosing.objects.filter(
                    organization=org,
                    closing_date__range=[compare_start, compare_end],
                )
                compare_sales_agg = compare_closings.aggregate(
                    card=Sum('actual_card'),
                    cash=Sum('actual_cash'),
                )
                compare_other = ClosingOtherSale.objects.filter(
                    closing__organization=org,
                    closing__closing_date__range=[compare_start, compare_end],
                ).aggregate(total=Sum('amount'))
                compare_total = (
                    float(compare_sales_agg['card'] or 0)
                    + float(compare_sales_agg['cash'] or 0)
                    + float(compare_other['total'] or 0)
                )
                compare_count = compare_closings.count()
                compare_avg = compare_total / max(compare_count, 1)

                impact_pct = (
                    round(((avg_daily - compare_avg) / compare_avg) * 100, 1)
                    if compare_avg > 0
                    else 0
                )

                # --- Staffing (holiday period) ---
                timesheets = Timesheet.objects.filter(
                    organization=org,
                    date__range=[h_start, h_end],
                    check_in__isnull=False,
                )
                staff_count = timesheets.values('user').distinct().count()
                total_shifts = timesheets.count()
                total_hours = sum(
                    t.worked_hours for t in timesheets if t.worked_hours
                )
                avg_staff_per_day = round(total_shifts / max(duration, 1), 1)
                splh = (
                    round(holiday_total / total_hours, 2)
                    if total_hours > 0
                    else 0
                )

                # --- Staffing (normal period) ---
                compare_ts = Timesheet.objects.filter(
                    organization=org,
                    date__range=[compare_start, compare_end],
                    check_in__isnull=False,
                )
                compare_shifts = compare_ts.count()
                normal_period_days = (compare_end - compare_start).days + 1
                normal_staff_avg = round(
                    compare_shifts / max(normal_period_days, 1), 1
                )
                compare_hours = sum(
                    t.worked_hours for t in compare_ts if t.worked_hours
                )
                normal_splh = (
                    round(compare_total / compare_hours, 2)
                    if compare_hours > 0
                    else 0
                )

                # Accumulate for summary
                total_holiday_sales += holiday_total
                total_holiday_days += days_with_data
                total_holiday_hours += total_hours
                total_normal_sales_for_avg += compare_total
                total_normal_days += compare_count

                holiday_results.append({
                    'id': h.id,
                    'name': h.name,
                    'name_ko': h.name_ko,
                    'category': h.category,
                    'start_date': str(h_start),
                    'end_date': str(h_end),
                    'impact': h.impact,
                    'sales': {
                        'total': round(holiday_total, 2),
                        'avg_daily': round(avg_daily, 2),
                        'days_with_data': days_with_data,
                        'normal_avg': round(compare_avg, 2),
                        'impact_pct': impact_pct,
                    },
                    'staffing': {
                        'staff_count': staff_count,
                        'total_shifts': total_shifts,
                        'total_hours': round(total_hours, 1),
                        'avg_staff_per_day': avg_staff_per_day,
                        'normal_staff_avg': normal_staff_avg,
                        'splh': splh,
                        'normal_splh': normal_splh,
                    },
                })

            # --- Yearly summary ---
            avg_holiday_daily = (
                round(total_holiday_sales / total_holiday_days, 2)
                if total_holiday_days > 0
                else 0
            )
            avg_normal_daily = (
                round(total_normal_sales_for_avg / total_normal_days, 2)
                if total_normal_days > 0
                else 0
            )
            overall_impact_pct = (
                round(
                    ((avg_holiday_daily - avg_normal_daily) / avg_normal_daily) * 100,
                    1,
                )
                if avg_normal_daily > 0
                else 0
            )
            avg_holiday_splh = (
                round(total_holiday_sales / total_holiday_hours, 2)
                if total_holiday_hours > 0
                else 0
            )

            return Response(
                {
                    'year': year,
                    'holidays': holiday_results,
                    'summary': {
                        'total_holiday_sales': round(total_holiday_sales, 2),
                        'total_holiday_days': total_holiday_days,
                        'avg_holiday_daily': avg_holiday_daily,
                        'avg_normal_daily': avg_normal_daily,
                        'overall_impact_pct': overall_impact_pct,
                        'total_holiday_hours': round(total_holiday_hours, 1),
                        'avg_holiday_splh': avg_holiday_splh,
                    },
                    'available_years': available_years,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class GeneratedReportViewSet(mixins.ListModelMixin,
                             mixins.RetrieveModelMixin,
                             viewsets.GenericViewSet):
    """
    생성된 리포트 조회 ViewSet
    - list: 생성된 리포트 목록
    - retrieve: 리포트 상세 정보
    """
    queryset = GeneratedReport.objects.all()
    serializer_class = GeneratedReportSerializer
    permission_classes = [IsAuthenticated, IsManager]
    filter_backends = [OrganizationFilterBackend]

    def get_queryset(self):
        """사용자의 조직에 해당하는 생성된 리포트만 조회"""
        queryset = super().get_queryset()
        return queryset.select_related('report', 'organization', 'generated_by__user')


class SkyReportViewSet(viewsets.ModelViewSet):
    """
    Sky Report — monthly financial report per store
    - list: all sky reports for the organization (filter by ?year=)
    - create: create a new monthly report
    - retrieve / update / partial_update / destroy
    - summary: yearly summary (PLAN vs ACTUAL style)
    """
    queryset = SkyReport.objects.all()
    serializer_class = SkyReportSerializer
    permission_classes = [IsAuthenticated, IsManager]
    filter_backends = [OrganizationFilterBackend]

    def get_queryset(self):
        qs = super().get_queryset()
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(year=year)
        return qs

    def _auto_calculate(self, instance):
        """Auto-calculate P&L fields from input fields."""
        import calendar

        d = Decimal
        # Total Sales & HQ Cash from garage inputs
        instance.total_sales_inc_gst = instance.total_sales_garage
        instance.hq_cash = instance.hq_cash_garage

        # EXCL GST = (Total Sales - HQ Cash) / 1.15 + HQ Cash
        excl_gst = ((instance.total_sales_inc_gst - instance.hq_cash) / d('1.15') + instance.hq_cash).quantize(d('0.01')) if instance.total_sales_inc_gst else d('0')

        # COGS (input is already excl.GST)
        instance.cogs = instance.total_cogs_xero if instance.total_cogs_xero else d('0')

        # Operating Expenses = (Total Expense - Labour - Sub)
        op_exp_excl = instance.total_expense_xero - instance.labour_xero - instance.sub_contractor_xero
        instance.operating_expenses = op_exp_excl

        # Wages = (labour / (payruns * 14)) * calendar_days
        days_in_month = calendar.monthrange(instance.year, instance.month)[1]
        payrun_days = instance.number_of_payruns * 14 if instance.number_of_payruns > 0 else 1
        if instance.labour_xero > 0:
            daily_rate = instance.labour_xero / d(str(payrun_days))
            instance.wages = (daily_rate * d(str(days_in_month))).quantize(d('0.01'))
        else:
            instance.wages = d('0')

        # Sub GST = sub_contractor * 1.15 * 3/23
        sub_inc_gst = instance.sub_contractor_xero * d('1.15')
        instance.sub_gst = (sub_inc_gst * d('3') / d('23')).quantize(d('0.01'))

        # Total labour cost = wages + sub_inc_gst - sub_gst
        total_labour = instance.wages + sub_inc_gst - instance.sub_gst

        # Payable GST = (sales - hq_cash - cogs_inc_gst - op_exp_inc_gst - sub_gst) * 3/23
        cogs_inc_gst = instance.total_cogs_xero * d('1.15') if instance.total_cogs_xero else d('0')
        gst_base = instance.total_sales_inc_gst - instance.hq_cash - cogs_inc_gst - (op_exp_excl * d('1.15')) - instance.sub_gst
        instance.payable_gst = (gst_base * d('3') / d('23')).quantize(d('0.01')) if gst_base > 0 else d('0')

        # Operating Profit = EXCL GST Sales - COGS - Operating Expenses - Total Labour
        instance.operating_profit = (excl_gst - instance.cogs - op_exp_excl - total_labour).quantize(d('0.01'))

        # Store total labour cost for display (sales_per_hour field repurposed)
        instance.sales_per_hour = total_labour.quantize(d('0.01'))
        # opening_sales_per_hour stores labour_xero for reference
        instance.opening_sales_per_hour = instance.labour_xero
        # other_sales = total_work_hours (manager input, don't overwrite)
        # tab_allowance_sales = opening_hours_per_day (manager input, don't overwrite)

        instance.save()

    def perform_create(self, serializer):
        from users.filters import get_target_org
        org = get_target_org(self.request)
        profile = self.request.user.profile
        instance = serializer.save(organization=org, created_by=profile)
        self._auto_calculate(instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._auto_calculate(instance)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Yearly summary — returns all months for a given year, grouped into halves."""
        from users.filters import get_target_org
        year = request.query_params.get('year', timezone.now().year)
        org = get_target_org(request)
        reports = SkyReport.objects.filter(organization=org, year=year).order_by('month')
        serializer = self.get_serializer(reports, many=True)

        # Build summary structure
        data = serializer.data
        first_half = [r for r in data if r['month'] <= 6]
        second_half = [r for r in data if r['month'] > 6]

        # Compute totals
        def compute_totals(reports_list):
            if not reports_list:
                return None
            total_sales = sum(float(r['total_sales_inc_gst']) for r in reports_list)
            total_cogs = sum(float(r['cogs']) for r in reports_list)
            total_wages = sum(float(r['wages']) for r in reports_list)
            return {
                'total_sales': total_sales,
                'total_cogs': total_cogs,
                'total_wages': total_wages,
                'cogs_ratio': round(total_cogs / total_sales * 100, 1) if total_sales else 0,
                'wage_ratio': round(total_wages / total_sales * 100, 1) if total_sales else 0,
            }

        return Response({
            'year': int(year),
            'first_half': first_half,
            'second_half': second_half,
            'first_half_totals': compute_totals(first_half),
            'second_half_totals': compute_totals(second_half),
            'annual_totals': compute_totals(data),
        })

    @action(detail=False, methods=['get'], url_path='auto-fill')
    def auto_fill(self, request):
        """Return auto-calculated values from DailyClosing for a given year/month."""
        import calendar
        from users.filters import get_target_org
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        org = get_target_org(request)

        days_in_month = calendar.monthrange(year, month)[1]
        start = date(year, month, 1)
        end = date(year, month, days_in_month)

        # If org is a company (has sub-stores), aggregate from all sub-stores
        target_orgs = org.all_stores
        closings = DailyClosing.objects.filter(
            organization__in=target_orgs,
            closing_date__range=[start, end],
        ).prefetch_related('other_sales', 'supplier_costs')

        total_sales = Decimal('0')
        hr_cash = Decimal('0')
        num_days = 0

        for c in closings:
            # Total Sales = pos_cash + pos_card + other_sales
            other_total = sum(o.amount for o in c.other_sales.all())
            day_total = c.pos_cash + c.pos_card + other_total
            total_sales += day_total

            # HQ Cash = HR Cash (actual_cash - bank_deposit), GST 없는 현금
            hr_cash += c.actual_cash - c.bank_deposit
            num_days += 1

        # Get opening hours from organization settings
        opening_hours = 0
        if org.opening_time and org.closing_time:
            from datetime import datetime as dt2, timedelta as td2
            open_dt = dt2.combine(date.today(), org.opening_time)
            close_dt = dt2.combine(date.today(), org.closing_time)
            if close_dt <= open_dt:
                close_dt += td2(days=1)  # e.g. 11:00 open, 21:00 close (next day handling)
            opening_hours = (close_dt - open_dt).seconds / 3600

        # Build sub-store info for frontend display
        sub_store_names = []
        if org.is_company:
            sub_store_names = list(target_orgs.values_list('name', flat=True))

        return Response({
            'total_sales_garage': float(total_sales),
            'hq_cash_garage': float(hr_cash),
            'number_of_days': num_days,
            'opening_hours_per_day': opening_hours,
            'is_company': org.is_company,
            'sub_store_names': sub_store_names,
        })

    def _aggregate_range(self, org, from_year, from_month, to_year, to_month):
        """Aggregate SkyReport data for a given year/month range."""
        from datetime import datetime as dt2, timedelta as td2
        from django.db.models import Q

        if from_year == to_year:
            q = Q(year=from_year, month__gte=from_month, month__lte=to_month)
        else:
            q = (
                Q(year=from_year, month__gte=from_month) |
                Q(year=to_year, month__lte=to_month) |
                Q(year__gt=from_year, year__lt=to_year)
            )

        reports = SkyReport.objects.filter(organization=org).filter(q).order_by('year', 'month')
        serializer = self.get_serializer(reports, many=True)
        data = serializer.data

        if not data:
            return {'reports': [], 'totals': {}}

        total_sales = sum(float(r.get('total_sales_inc_gst', 0) or 0) for r in data)
        excl_gst = sum(float(r.get('excl_gst_sales', 0) or 0) for r in data)
        cogs = sum(float(r.get('cogs', 0) or 0) for r in data)
        op_exp = sum(float(r.get('operating_expenses', 0) or 0) for r in data)
        wages = sum(float(r.get('wages', 0) or 0) for r in data)
        labour = sum(float(r.get('sales_per_hour', 0) or 0) for r in data)  # total_labour field
        profit = sum(float(r.get('operating_profit', 0) or 0) for r in data)
        days = sum(int(r.get('number_of_days', 0) or 0) for r in data)
        hq_cash = sum(float(r.get('hq_cash', 0) or 0) for r in data)

        # Opening hours from org settings
        opening_hours_per_day = 0
        if org.opening_time and org.closing_time:
            open_dt = dt2.combine(date.today(), org.opening_time)
            close_dt = dt2.combine(date.today(), org.closing_time)
            if close_dt <= open_dt:
                close_dt += td2(days=1)
            opening_hours_per_day = (close_dt - open_dt).seconds / 3600

        total_opening_hours = days * opening_hours_per_day
        total_work_hours = sum(float(r.get('other_sales', 0) or 0) for r in data)
        total_tabs = sum(float(r.get('pos_sales', 0) or 0) for r in data)

        return {
            'reports': data,
            'totals': {
                'total_sales': total_sales,
                'excl_gst': excl_gst,
                'cogs': cogs,
                'cogs_ratio': round(cogs / excl_gst * 100, 1) if excl_gst else 0,
                'op_exp': op_exp,
                'wages': wages,
                'labour': labour,
                'labour_ratio': round(labour / excl_gst * 100, 1) if excl_gst else 0,
                'profit': profit,
                'profit_ratio': round(profit / excl_gst * 100, 1) if excl_gst else 0,
                'days': days,
                'hq_cash': hq_cash,
                'sales_per_day': round(excl_gst / days, 2) if days else 0,
                'sales_per_tab': round(excl_gst / total_tabs, 2) if total_tabs else 0,
                'sales_per_labour_hr': round(excl_gst / total_work_hours, 2) if total_work_hours else 0,
                'sales_per_opening_hr': round(excl_gst / total_opening_hours, 2) if total_opening_hours else 0,
            },
        }

    @action(detail=False, methods=['get'], url_path='range-summary')
    def range_summary(self, request):
        """Aggregate Sky Reports for a date range with multi-year comparison."""
        from users.filters import get_target_org
        org = get_target_org(request)

        from_year = int(request.query_params.get('from_year', timezone.now().year))
        from_month = int(request.query_params.get('from_month', 1))
        to_year = int(request.query_params.get('to_year', from_year))
        to_month = int(request.query_params.get('to_month', 12))

        # Current period
        current = self._aggregate_range(org, from_year, from_month, to_year, to_month)

        # Previous year (same months, 1 year back)
        prev_1 = self._aggregate_range(org, from_year - 1, from_month, to_year - 1, to_month)
        prev_2 = self._aggregate_range(org, from_year - 2, from_month, to_year - 2, to_month)

        def calc_yoy(curr_totals, prev_totals):
            if not prev_totals or not curr_totals:
                return None
            result = {}
            for key in ['total_sales', 'cogs', 'labour', 'profit', 'excl_gst']:
                cv = curr_totals.get(key, 0) or 0
                pv = prev_totals.get(key, 0) or 0
                result[key] = round((cv - pv) / pv * 100, 1) if pv else 0
            return result

        ct = current.get('totals', {})
        p1t = prev_1.get('totals', {}) if prev_1.get('reports') else None
        p2t = prev_2.get('totals', {}) if prev_2.get('reports') else None

        comparison = {}
        if p1t:
            comparison['prev_1'] = {'totals': p1t, 'reports': prev_1['reports']}
            comparison['yoy_1'] = calc_yoy(ct, p1t)
        if p2t:
            comparison['prev_2'] = {'totals': p2t, 'reports': prev_2['reports']}
            comparison['yoy_2'] = calc_yoy(p1t, p2t) if p1t else calc_yoy(ct, p2t)

        return Response({
            'from_year': from_year,
            'from_month': from_month,
            'to_year': to_year,
            'to_month': to_month,
            'months_count': len(current.get('reports', [])),
            'reports': current.get('reports', []),
            'totals': ct,
            'comparison': comparison,
        })

    @action(detail=False, methods=['get'], url_path='ai-analysis')
    def ai_analysis(self, request):
        """Generate AI-powered P&L analysis for a date range."""
        from users.filters import get_target_org
        from .services import generate_sky_report_analysis

        org = get_target_org(request)
        from_year = int(request.query_params.get('from_year', timezone.now().year))
        from_month = int(request.query_params.get('from_month', 1))
        to_year = int(request.query_params.get('to_year', from_year))
        to_month = int(request.query_params.get('to_month', 12))

        # Build range data with comparison
        current = self._aggregate_range(org, from_year, from_month, to_year, to_month)
        prev_1 = self._aggregate_range(org, from_year - 1, from_month, to_year - 1, to_month)
        prev_2 = self._aggregate_range(org, from_year - 2, from_month, to_year - 2, to_month)

        ct = current.get('totals', {})
        p1t = prev_1.get('totals', {}) if prev_1.get('reports') else None
        p2t = prev_2.get('totals', {}) if prev_2.get('reports') else None

        def calc_yoy(curr_totals, prev_totals):
            if not prev_totals or not curr_totals:
                return None
            result = {}
            for key in ['total_sales', 'cogs', 'labour', 'profit']:
                cv = curr_totals.get(key, 0) or 0
                pv = prev_totals.get(key, 0) or 0
                result[key.replace('total_', '')] = round((cv - pv) / pv * 100, 1) if pv else 0
            return result

        comparison = {}
        if p1t:
            comparison['prev_1'] = {'totals': p1t}
            comparison['yoy_1'] = calc_yoy(ct, p1t)
        if p2t:
            comparison['prev_2'] = {'totals': p2t}
            comparison['yoy_2'] = calc_yoy(p1t, p2t) if p1t else None

        range_data = {
            'from_year': from_year,
            'from_month': from_month,
            'to_year': to_year,
            'to_month': to_month,
            'reports': current.get('reports', []),
            'totals': ct,
            'comparison': comparison,
        }

        result = generate_sky_report_analysis(range_data, org.name)
        if result is None:
            return Response({'error': 'AI analysis unavailable'}, status=503)

        return Response(result)

    @action(detail=False, methods=['get'])
    def template(self, request):
        """Download an Excel template for bulk Sky Report data entry."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sky Report Data'

        # Styles
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        section_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
        section_font = Font(bold=True, size=10, color='1E40AF')
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )

        # Column headers: Year, Month, then all data fields
        columns = [
            ('Year', 10), ('Month (1-12)', 14),
            # Main financial
            ('Total Sales (GST incl.)', 22), ('HQ CASH', 14), ('Pos Sales', 14), ('Other Sales', 14),
            ('COGS', 14), ('Operating Expenses', 20), ('Wages & Salaries', 18),
            ('Sales per Hour', 16), ('Opening Sales/Hr', 18), ('Tab Allowance Sales', 20),
            ('Payable GST', 14), ('Sub GST', 12), ('Operating Profit', 18),
            # External
            ('Total Sales (Garage)', 22), ('HQ CASH (Garage)', 18),
            ('Total COGS (Xero)', 18), ('Total Expense (Xero)', 22),
            ('Labour (Xero)', 16), ('Sub-contractor (Xero)', 22),
            ('Number of Days', 16), ('Number of Payruns', 18),
            # Goals
            ('Sales Goal', 14), ('COGS Goal', 12), ('Wage Goal', 12),
            ('Review Rating', 14), ('Review Goal', 14),
            # Hygiene
            ('Hygiene Grade', 14),
            # Notes
            ('Sales Notes', 20), ('COGS Notes', 20), ('Wage Notes', 20), ('Next Month Notes', 20),
        ]

        # Write header row
        for col_idx, (name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Add sample row
        sample = [
            2025, 4,
            179317.90, 0, 179317.90, 0,
            52962.28, 0, 37175.19,
            0, 0, 0,
            0, 0, 89180.43,
            0, 0,
            52962.28, 0,
            37175.19, 0,
            30, 2,
            200000, 50000, 40000,
            4.7, 4.8,
            'A',
            '', '', '', '',
        ]
        for col_idx, val in enumerate(sample, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx <= 2:
                cell.fill = section_fill
                cell.font = section_font

        # Add empty rows for data entry (12 months)
        for row_idx in range(3, 15):
            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

        # Instructions sheet
        ws2 = wb.create_sheet('Instructions')
        instructions = [
            ['Sky Report Template — Instructions'],
            [''],
            ['1. Fill in data in the "Sky Report Data" sheet'],
            ['2. Year: Enter the year (e.g., 2025)'],
            ['3. Month: Enter 1-12 (1=January, 12=December)'],
            ['4. All monetary values are in NZD'],
            ['5. Hygiene Grade: A, B, C, D, or E'],
            ['6. You can add as many rows as needed'],
            ['7. The sample row (row 2) can be deleted or overwritten'],
            ['8. Upload this file on the Sky Report page'],
            [''],
            ['Field Mapping (Korean):'],
            ['Total Sales (GST incl.) = 총매출 (GST 포함)'],
            ['HQ CASH = 현금'],
            ['Pos Sales = 포스 매출'],
            ['COGS = 매출원가'],
            ['Operating Expenses = 운영비용'],
            ['Wages & Salaries = 인건비'],
            ['Operating Profit = 영업이익 (세전)'],
        ]
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx >= 12:
                    cell.font = Font(size=10, color='666666')
        ws2.column_dimensions['A'].width = 50

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sky_report_template.xlsx"'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def upload(self, request):
        """Upload an Excel file to bulk-create Sky Report entries."""
        import openpyxl

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': 'Only .xlsx files are supported'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            org = request.user.profile.organization
            profile = request.user.profile

            # Field mapping: column index (0-based after Year,Month) → model field
            field_map = [
                'total_sales_inc_gst', 'hq_cash', 'pos_sales', 'other_sales',
                'cogs', 'operating_expenses', 'wages',
                'sales_per_hour', 'opening_sales_per_hour', 'tab_allowance_sales',
                'payable_gst', 'sub_gst', 'operating_profit',
                'total_sales_garage', 'hq_cash_garage',
                'total_cogs_xero', 'total_expense_xero',
                'labour_xero', 'sub_contractor_xero',
                'number_of_days', 'number_of_payruns',
                'sales_goal', 'cogs_goal', 'wage_goal',
                'review_rating', 'review_goal',
                'hygiene_grade',
                'sales_notes', 'cogs_notes', 'wage_notes', 'next_month_notes',
            ]

            created = 0
            updated = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not row[0] or not row[1]:
                    continue

                try:
                    year = int(row[0])
                    month = int(row[1])
                    if month < 1 or month > 12:
                        errors.append(f'Row {row_idx}: Invalid month {month}')
                        continue
                except (ValueError, TypeError):
                    errors.append(f'Row {row_idx}: Invalid year/month')
                    continue

                data = {}
                for i, field_name in enumerate(field_map):
                    val = row[i + 2] if i + 2 < len(row) else None

                    if field_name in ('hygiene_grade', 'sales_notes', 'cogs_notes', 'wage_notes', 'next_month_notes'):
                        data[field_name] = str(val).strip() if val else ''
                    elif field_name in ('number_of_days', 'number_of_payruns'):
                        try:
                            data[field_name] = int(val) if val else 0
                        except (ValueError, TypeError):
                            data[field_name] = 0
                    else:
                        try:
                            data[field_name] = Decimal(str(val)) if val else Decimal('0')
                        except Exception:
                            data[field_name] = Decimal('0')

                obj, was_created = SkyReport.objects.update_or_create(
                    organization=org,
                    year=year,
                    month=month,
                    defaults={**data, 'created_by': profile},
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

            return Response({
                'created': created,
                'updated': updated,
                'errors': errors,
                'total': created + updated,
            })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StoreEvaluationViewSet(viewsets.ModelViewSet):
    """
    Store Evaluation — semi-annual manager performance evaluation
    - list: all evaluations for the organization (filter by ?year=)
    - create/update: CEO/HQ only
    - auto_fill: pull achievements from SkyReport data
    - toggle_lock: lock/unlock evaluation (CEO/HQ only)
    """
    queryset = StoreEvaluation.objects.all()
    serializer_class = StoreEvaluationSerializer
    permission_classes = [IsAuthenticated, IsManager]
    filter_backends = [OrganizationFilterBackend]

    def get_queryset(self):
        qs = super().get_queryset()
        year = self.request.query_params.get('year')
        period = self.request.query_params.get('period_type')
        if year:
            qs = qs.filter(year=year)
        if period:
            qs = qs.filter(period_type=period)
        return qs.select_related('organization', 'created_by__user')

    def _check_ceo_hq(self, request):
        """Check if user is CEO or HQ."""
        return request.user.profile.role in ['CEO', 'HQ']

    def _calculate_scores(self, instance):
        """Calculate all evaluation scores based on targets vs achievements."""
        d = Decimal

        # --- Guarantee amount ---
        if instance.guarantee_pct and instance.net_profit:
            instance.guarantee_amount = (instance.net_profit * instance.guarantee_pct / d('100')).quantize(d('0.01'))
        else:
            instance.guarantee_amount = d('0')

        # --- Incentive pool ---
        if instance.incentive_amount > 0:
            instance.incentive_pool = instance.incentive_amount
        elif instance.incentive_pct > 0 and instance.net_profit > 0:
            instance.incentive_pool = (instance.net_profit * instance.incentive_pct / d('100')).quantize(d('0.01'))
        else:
            instance.incentive_pool = d('0')

        # --- Sales score (max 25) ---
        if instance.sales_target and instance.sales_target > 0:
            ratio = float(instance.sales_achievement / instance.sales_target * 100)
            if ratio >= 110:
                instance.sales_score = 25
            elif ratio >= 100:
                instance.sales_score = 15
            elif ratio >= 90:
                instance.sales_score = 10
            else:
                instance.sales_score = 0
        else:
            instance.sales_score = 0

        # --- COGS score (max 20) ---
        # diff in percentage points (achievement - target); positive = worse
        if instance.cogs_target > 0:
            diff = float(instance.cogs_achievement - instance.cogs_target) * 100  # convert ratio to %p
            if diff <= 0:
                instance.cogs_score = 20
            elif diff <= 1:
                instance.cogs_score = 18
            elif diff <= 2:
                instance.cogs_score = 14
            else:
                instance.cogs_score = 7
        else:
            instance.cogs_score = 0

        # --- Wage score (max 25) ---
        if instance.wage_target > 0:
            diff = float(instance.wage_achievement - instance.wage_target) * 100  # convert ratio to %p
            if diff <= 0:
                instance.wage_score = 25
            elif diff <= 1:
                instance.wage_score = 15
            elif diff <= 2:
                instance.wage_score = 10
            else:
                instance.wage_score = 0
        else:
            instance.wage_score = 0

        # --- Service score (max 10) ---
        rating = float(instance.service_rating)
        if rating >= 4.8:
            instance.service_score = 10
        elif rating >= 4.5:
            instance.service_score = 7
        elif rating >= 4.2:
            instance.service_score = 5
        else:
            instance.service_score = 0

        # --- Hygiene score (max 10) ---
        if instance.hygiene_months >= 18:
            instance.hygiene_score = 10
        elif instance.hygiene_months >= 12:
            instance.hygiene_score = 5
        else:
            instance.hygiene_score = 0

        # --- Leadership score points (max 10) ---
        ls = instance.leadership_score
        if ls >= 5:
            instance.leadership_score_points = 10
        elif ls >= 4:
            instance.leadership_score_points = 5
        elif ls >= 3:
            instance.leadership_score_points = 2
        else:
            instance.leadership_score_points = 0

        # --- Total ---
        instance.total_score = (
            instance.sales_score + instance.cogs_score + instance.wage_score +
            instance.service_score + instance.hygiene_score + instance.leadership_score_points
        )
        instance.payout_ratio = d(str(instance.total_score)) / d('100')

        instance.save()

    def perform_create(self, serializer):
        from users.filters import get_target_org
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can create evaluations.')
        org = get_target_org(self.request)
        profile = self.request.user.profile
        instance = serializer.save(organization=org, created_by=profile)
        self._calculate_scores(instance)

    def perform_update(self, serializer):
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can edit evaluations.')
        instance = serializer.instance
        if instance.is_locked:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('This evaluation is locked and cannot be modified.')
        instance = serializer.save()
        self._calculate_scores(instance)

    def perform_destroy(self, instance):
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can delete evaluations.')
        if instance.is_locked:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('This evaluation is locked and cannot be deleted.')
        instance.delete()

    @action(detail=False, methods=['get'], url_path='auto-fill')
    def auto_fill(self, request):
        """Auto-fill achievements from SkyReport data for the evaluation period."""
        from users.filters import get_target_org
        from django.db.models import Q

        org = get_target_org(request)
        year = int(request.query_params.get('year', timezone.now().year))
        period_type = request.query_params.get('period_type', 'H1')

        # Determine month range: H1=Apr-Sep, H2=Oct-Mar (crosses year boundary)
        if period_type == 'H1':
            sky_reports = SkyReport.objects.filter(
                organization=org, year=year, month__in=range(4, 10),
            )
        else:
            sky_reports = SkyReport.objects.filter(
                organization=org,
            ).filter(
                Q(year=year, month__in=[10, 11, 12]) |
                Q(year=year + 1, month__in=[1, 2, 3])
            )

        if not sky_reports.exists():
            return Response({'error': 'No SkyReport data found for this period.'}, status=status.HTTP_404_NOT_FOUND)

        total_sales = sky_reports.aggregate(total=Sum('total_sales_inc_gst'))['total'] or Decimal('0')
        total_cogs = sky_reports.aggregate(total=Sum('cogs'))['total'] or Decimal('0')
        total_wages = sky_reports.aggregate(total=Sum('wages'))['total'] or Decimal('0')
        excl_gst = sky_reports.aggregate(total=Sum('excl_gst_sales'))['total'] or Decimal('0')

        cogs_pct = (total_cogs / excl_gst).quantize(Decimal('0.0001')) if excl_gst else Decimal('0')
        wage_pct = (total_wages / excl_gst).quantize(Decimal('0.0001')) if excl_gst else Decimal('0')

        latest_report = sky_reports.order_by('-year', '-month').first()
        service_rating = float(latest_report.review_rating) if latest_report and latest_report.review_rating > 0 else 0

        return Response({
            'total_sales': float(total_sales),
            'cogs_percent': float(cogs_pct),
            'wage_percent': float(wage_pct),
            'service_rating': service_rating,
            'months_found': sky_reports.count(),
        })

    @action(detail=True, methods=['post'])
    def toggle_lock(self, request, pk=None):
        """Toggle lock status of an evaluation. CEO/HQ only."""
        if not self._check_ceo_hq(request):
            return Response({'error': 'Only CEO/HQ can lock/unlock evaluations.'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        instance.is_locked = not instance.is_locked
        instance.save()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def history(self, request):
        """Return all evaluations for a store for the history table."""
        from users.filters import get_target_org
        org = get_target_org(request)
        if not org:
            return Response({'error': 'No store selected.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = StoreEvaluation.objects.filter(organization=org).order_by('-year', '-period_type')
        results = []
        for ev in qs:
            results.append({
                'id': ev.id,
                'year': ev.year,
                'period_type': ev.period_type,
                'total_score': ev.total_score,
                'sales_score': ev.sales_score,
                'cogs_score': ev.cogs_score,
                'wage_score': ev.wage_score,
                'service_score': ev.service_score,
                'hygiene_score': ev.hygiene_score,
                'leadership_score_points': ev.leadership_score_points,
                'is_locked': ev.is_locked,
            })
        return Response(results)


class ProfitShareViewSet(viewsets.ModelViewSet):
    """
    Partner Profit Share — semi-annual profit distribution per store
    - list: all profit shares (filter by ?year=&period_type=)
    - create/update: CEO/HQ only, with nested partners
    - toggle_lock: lock/unlock (CEO/HQ only)
    - auto_calculate: recalculate all partner amounts from percentages
    """
    queryset = ProfitShare.objects.all()
    serializer_class = ProfitShareSerializer
    permission_classes = [IsAuthenticated, IsManager]
    filter_backends = [OrganizationFilterBackend]

    def get_queryset(self):
        qs = super().get_queryset()
        year = self.request.query_params.get('year')
        period = self.request.query_params.get('period_type')
        if year:
            qs = qs.filter(year=year)
        if period:
            qs = qs.filter(period_type=period)
        return qs.select_related('organization', 'created_by__user').prefetch_related('partners')

    def _check_ceo_hq(self, request):
        return request.user.profile.role in ['CEO', 'HQ']

    def perform_create(self, serializer):
        from users.filters import get_target_org
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can create profit shares.')
        org = get_target_org(self.request)
        profile = self.request.user.profile
        serializer.save(organization=org, created_by=profile)

    def perform_update(self, serializer):
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can edit profit shares.')
        instance = serializer.instance
        if instance.is_locked:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('This profit share is locked and cannot be modified.')
        serializer.save()

    def perform_destroy(self, instance):
        if not self._check_ceo_hq(self.request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Only CEO/HQ can delete profit shares.')
        if instance.is_locked:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('This profit share is locked and cannot be deleted.')
        instance.delete()

    @action(detail=True, methods=['post'])
    def auto_calculate(self, request, pk=None):
        """Recalculate all partner amounts from percentages."""
        instance = self.get_object()
        if instance.is_locked:
            return Response({'error': 'Profit share is locked.'}, status=status.HTTP_403_FORBIDDEN)
        if not self._check_ceo_hq(request):
            return Response({'error': 'Only CEO/HQ can recalculate.'}, status=status.HTTP_403_FORBIDDEN)

        instance.calculate_totals()
        instance.save()

        # Calculate non-owner, non-fixed partners first
        for partner in instance.partners.exclude(partner_type='OWNER').filter(fixed_amount=0):
            partner.calculate_amounts()
            partner.save()

        # Calculate fixed-amount partners
        for partner in instance.partners.filter(fixed_amount__gt=0):
            partner.calculate_amounts()
            partner.save()

        # Calculate owner last (gets remainder)
        for partner in instance.partners.filter(partner_type='OWNER'):
            partner.calculate_amounts()
            partner.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def history(self, request):
        """Return all profit shares for a store (all years/periods) for the history dashboard."""
        from users.filters import get_target_org
        org = get_target_org(request)
        if not org:
            return Response({'error': 'No store selected.'}, status=status.HTTP_400_BAD_REQUEST)

        qs = ProfitShare.objects.filter(organization=org).order_by('year', 'period_type')
        results = []
        for ps in qs.prefetch_related('partners'):
            net_profit_total = float(ps.net_profit_account + ps.net_profit_cash)
            incentive_total = float(ps.incentive_account + ps.incentive_cash)
            results.append({
                'id': ps.id,
                'year': ps.year,
                'period_type': ps.period_type,
                'period_display': ps.get_period_type_display(),
                'net_profit_total': net_profit_total,
                'incentive_total': incentive_total,
                'partner_count': ps.partners.count(),
                'is_locked': ps.is_locked,
            })
        return Response(results)

    @action(detail=True, methods=['post'])
    def toggle_lock(self, request, pk=None):
        """Toggle lock status. CEO/HQ only.
        When locking: auto-create CQ transaction records for each partner's distribution.
        When unlocking: delete auto-created CQ records for this profit share.
        """
        if not self._check_ceo_hq(request):
            return Response({'error': 'Only CEO/HQ can lock/unlock profit shares.'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        instance.is_locked = not instance.is_locked
        instance.save()

        from closing.models import CQTransaction

        if instance.is_locked:
            # Locking: create CQ transaction records for each partner
            # Determine period end date and period label
            if instance.period_type == 'H1':
                # H1 = Apr-Sep, end date = Sep 30
                period_end_date = date(instance.year, 9, 30)
                period_label = f"{instance.year}-Oct"
            else:
                # H2 = Oct-Mar, end date = Mar 31 of next year
                period_end_date = date(instance.year + 1, 3, 31)
                period_label = f"{instance.year + 1}-Apr"

            store_name = instance.organization.name if instance.organization else ''
            profile = request.user.profile

            for partner in instance.partners.all():
                # OWNER → COLLECTION (owner's profit)
                # NON_EQUITY → INCENTIVE (partner incentive)
                # EQUITY → PROFIT (equity distribution)
                if partner.partner_type == 'OWNER':
                    tx_type = 'COLLECTION'
                    type_label = 'Owner Profit'
                elif partner.partner_type == 'NON_EQUITY':
                    tx_type = 'INCENTIVE'
                    type_label = 'Incentive'
                else:
                    tx_type = 'PROFIT'
                    type_label = 'Equity Share'

                # Single record per partner (Account + Cash combined)
                total_amount = (partner.total_account or Decimal('0')) + (partner.total_cash or Decimal('0'))
                if total_amount and total_amount != Decimal('0'):
                    CQTransaction.objects.create(
                        organization=instance.organization,
                        date=period_end_date,
                        store_name=store_name,
                        transaction_type=tx_type,
                        person=partner.name,
                        amount=total_amount,
                        account_type='ACCOUNT',
                        note=f"{type_label} {instance.year} {instance.get_period_type_display()} - {partner.name}",
                        period=period_label,
                        profit_share=instance,
                        created_by=profile,
                    )
        else:
            # Unlocking: delete auto-created CQ records for this profit share
            CQTransaction.objects.filter(profit_share=instance).delete()

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='pull-score')
    def pull_score(self, request):
        """Pull evaluation score for the same period."""
        from users.filters import get_target_org
        org = get_target_org(request)
        year = request.query_params.get('year')
        period_type = request.query_params.get('period_type')

        if not year or not period_type:
            return Response({'error': 'year and period_type are required.'}, status=status.HTTP_400_BAD_REQUEST)

        evaluation = StoreEvaluation.objects.filter(
            organization=org, year=int(year), period_type=period_type
        ).first()

        if not evaluation:
            return Response({'error': 'No evaluation found for this period.'}, status=status.HTTP_404_NOT_FOUND)

        total = evaluation.total_score or 0
        return Response({
            'total_score': total,
            'score_percentage': float(total) / 100.0 if total else 0,
            'sales_score': evaluation.sales_score,
            'cogs_score': evaluation.cogs_score,
            'wage_score': evaluation.wage_score,
            'service_score': evaluation.service_score,
            'hygiene_score': evaluation.hygiene_score,
            'leadership_score_points': evaluation.leadership_score_points,
        })
