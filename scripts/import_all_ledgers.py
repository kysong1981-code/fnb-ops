"""
Import QT, ChCh, KRW ledger data from Excel into CQ Transactions.
Run on server: docker exec fnb-ops-backend python /app/scripts/import_all_ledgers.py
"""
import os, sys, django
sys.path.insert(0, '/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from decimal import Decimal
from datetime import date, datetime
from closing.models import CQTransaction
from users.models import Organization
import openpyxl

EXCEL_PATH = '/app/scripts/1.xlsx'
ORG = Organization.objects.get(id=3)  # OneOps

def parse_date_cell(val, current_date, current_year):
    """Parse various date formats from Excel cells."""
    if val is None:
        return current_date
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    s = str(val).strip()
    if not s or s == '0':
        return current_date

    # "16.04.24" format (DD.MM.YY)
    if '.' in s and len(s.split('.')) == 3:
        parts = s.split('.')
        try:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return date(y, m, d)
        except:
            pass

    # "10월", "4월", "11월", etc.
    if '월' in s:
        s_clean = s.replace('월', '').strip()
        # Handle "11월21일" or "11월 4일"
        if '일' in s_clean:
            parts = s_clean.split('일')[0]
            if ' ' in parts:
                month_part, day_part = parts.rsplit(' ', 1)
            else:
                # Could be like "21" after removing 일
                # The month was before 월
                month_part = s.split('월')[0].strip()
                day_part = s.split('월')[1].replace('일', '').strip()
            try:
                m = int(month_part)
                d = int(day_part) if day_part else 1
                return date(current_year, m, d)
            except:
                pass
        try:
            m = int(s_clean)
            # If month < current month and we're in a sequence, might be next year
            return date(current_year, m, 1)
        except:
            pass

    # Just a number like "4", "5", "17", "27", "28" - day of month
    try:
        day_num = int(s)
        if 1 <= day_num <= 31:
            return date(current_date.year, current_date.month, min(day_num, 28))
    except:
        pass

    # "Total" or other text - skip
    if s.lower() == 'total':
        return current_date

    return current_date


def import_sheet(ws, person, opening_balance, start_row, currency='NZD'):
    """Import a sheet into CQ Transactions."""
    print(f"\n{'='*50}")
    print(f"Importing {person} (Opening: {opening_balance}, Currency: {currency})")
    print(f"{'='*50}")

    entries = []

    # Opening balance
    entries.append({
        'date': date(2024, 1, 1),
        'type': 'BALANCE',
        'amount': Decimal(str(opening_balance)),
        'store': '',
        'note': f'Opening Balance 2024',
        'period': '2024-Jan',
    })

    current_date = date(2024, 1, 1)
    current_year = 2024
    period = '2024-H1'

    for row_idx in range(start_row, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value  # Date
        b = ws.cell(row=row_idx, column=2).value  # Income
        c = ws.cell(row=row_idx, column=3).value  # Expense
        d = ws.cell(row=row_idx, column=4).value  # Note
        e = ws.cell(row=row_idx, column=5).value  # Balance (for verification)

        income = float(b) if b and b != 0 else 0
        expense = float(c) if c and c != 0 else 0
        note = str(d).strip() if d else ''

        # Skip header rows, empty rows, rows with 0/0
        if income == 0 and expense == 0:
            # Check if it's a date-only marker
            if a is not None:
                new_date = parse_date_cell(a, current_date, current_year)
                if new_date != current_date:
                    current_date = new_date
                    current_year = current_date.year
                    # Update period based on month
                    if current_date.month >= 10:
                        period = f'{current_date.year}-Oct'
                    elif current_date.month >= 4:
                        period = f'{current_date.year}-Apr'
                    else:
                        period = f'{current_date.year}-Jan'
            continue

        # Parse date
        if a is not None:
            new_date = parse_date_cell(a, current_date, current_year)
            if new_date != current_date:
                current_date = new_date
                current_year = current_date.year
                if current_date.month >= 10:
                    period = f'{current_date.year}-Oct'
                elif current_date.month >= 4:
                    period = f'{current_date.year}-Apr'
                else:
                    period = f'{current_date.year}-Jan'

        # Determine store name from note for income entries
        store = ''
        if income > 0:
            store_map = {
                'q airport': 'Q Airport', 'q airport': 'Q Airport',
                '5mile': '5 Mile', '5 mile': '5 Mile',
                'teppan': 'Teppan',
                't1': 'T1', 't2': 'T2',
                'mart': 'Mart',
                'f thai': 'F Thai',
                'a auckland': 'A Auckland', 'air auckland': 'A Auckland',
                'takanini': 'Takanini',
                'mums': 'Mums',
                'manawabay': 'Manawa Bay', 'manawa': 'Manawa Bay',
                'a chch': 'A ChCh', 'chch airport': 'ChCh Airport',
                'riverside': 'Riverside',
                '더니든': 'Dunedin',
                '프랜크톤': 'Frankton Thai',
                'serena': 'Q Airport',
                '치치공항': 'ChCh Airport',
                '리버사이드': 'Riverside',
                '마나와': 'Manawa Bay',
            }
            note_lower = note.lower()
            for key, val in store_map.items():
                if key in note_lower:
                    store = val
                    break

        # Both income and expense same amount = pass-through (net zero)
        if income > 0 and expense > 0:
            if income == expense:
                # Net zero - record both
                entries.append({
                    'date': current_date,
                    'type': 'TRANSFER',
                    'amount': Decimal(str(income)),
                    'store': store or note,
                    'note': note,
                    'period': period,
                })
                entries.append({
                    'date': current_date,
                    'type': 'EXPENSE',
                    'amount': Decimal(str(expense)),
                    'store': '',
                    'note': f'{note} (pass-through)',
                    'period': period,
                })
            else:
                # Different amounts - record separately
                entries.append({
                    'date': current_date,
                    'type': 'TRANSFER',
                    'amount': Decimal(str(income)),
                    'store': store or note,
                    'note': note,
                    'period': period,
                })
                entries.append({
                    'date': current_date,
                    'type': 'EXPENSE',
                    'amount': Decimal(str(expense)),
                    'store': '',
                    'note': note,
                    'period': period,
                })
        elif income > 0:
            entries.append({
                'date': current_date,
                'type': 'TRANSFER',
                'amount': Decimal(str(income)),
                'store': store or note,
                'note': note,
                'period': period,
            })
        elif expense > 0:
            entries.append({
                'date': current_date,
                'type': 'EXPENSE',
                'amount': Decimal(str(expense)),
                'store': '',
                'note': note,
                'period': period,
            })

        # Advance date slightly for ordering (same-date entries)
        # We keep same date but rely on creation order

    # Create CQ Transactions
    created = 0
    for e in entries:
        CQTransaction.objects.create(
            organization=ORG,
            date=e['date'],
            store_name=e['store'],
            transaction_type=e['type'],
            person=person,
            amount=e['amount'],
            account_type='CASH',
            note=e['note'],
            period=e['period'],
        )
        created += 1

    # Verify balance
    INFLOW = ('COLLECTION', 'BALANCE', 'TRANSFER')
    qs = CQTransaction.objects.filter(person=person).order_by('date', 'id')
    balance = Decimal('0')
    for t in qs:
        if t.transaction_type in INFLOW:
            balance += t.amount
        else:
            balance -= t.amount

    print(f"  Created: {created} transactions")
    print(f"  Final balance: {balance:,.2f}")
    return created


# ==================================================
# Main
# ==================================================
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Delete existing QT/ChCh/KRW transactions to avoid duplicates
for p in ['QT', 'ChCh', 'KRW']:
    deleted = CQTransaction.objects.filter(person=p).delete()
    print(f"Deleted existing {p} transactions: {deleted}")

total = 0

# QT: Opening $21,324.80, data starts at row 3
total += import_sheet(wb['QT'], 'QT', 21324.80, start_row=3)

# ChCh: Opening $2,886.50, data starts at row 8
total += import_sheet(wb['chch'], 'ChCh', 2886.50, start_row=8)

# KRW: Opening ₩10,348,277, data starts at row 3
total += import_sheet(wb['Sheet3'], 'KRW', 10348277, start_row=3, currency='KRW')

print(f"\n{'='*50}")
print(f"TOTAL: {total} transactions imported")
print(f"{'='*50}")
