"""
Import ChCh KRW Account Ledger from Excel into CQ Transactions.

Usage:
  1. Copy the Excel into the container:
     docker cp ~/Downloads/통합\ 문서21.xlsx fnb-ops-backend:/tmp/chch_ledger.xlsx
  2. Copy this script into the container:
     docker cp scripts/import_chch_ledger.py fnb-ops-backend:/tmp/import_chch_ledger.py
  3. Run:
     docker exec fnb-ops-backend python /tmp/import_chch_ledger.py

The Excel has columns: Date (A), Income (B), Expense (C), Note (D), Balance (E)
Row 1 = opening balance header, Row 2 = column headers, Row 3+ = data.
All amounts are in KRW. account_type = 'KRW', person = 'ChCh'.
"""
import os, sys, django
sys.path.insert(0, '/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from decimal import Decimal
from datetime import date, datetime
from closing.models import CQTransaction
from users.models import Organization

# ============================================================
# Config
# ============================================================
PERSON = 'ChCh'
ACCT = 'KRW'
ORG_ID = 1  # T2 org

org = Organization.objects.get(id=ORG_ID)
print(f"Organization: {org.name} (id={org.id})")

# ============================================================
# Step 1: Delete existing ChCh transactions
# ============================================================
existing = CQTransaction.objects.filter(person=PERSON)
count = existing.count()
print(f"Deleting {count} existing CQTransaction records for person='{PERSON}'...")
existing.delete()
print(f"Deleted.")

# ============================================================
# Step 2: Parse Excel data
# ============================================================
import openpyxl

EXCEL_PATH = '/tmp/chch_ledger.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Sheet1']

# Row 1: opening balance
opening_balance = ws.cell(1, 5).value  # E1 = 10348277
print(f"Opening balance from Excel: {opening_balance}")

entries = []


def make_entry(dt, typ, amount, note='', period=''):
    return {
        'date': dt,
        'transaction_type': typ,
        'amount': Decimal(str(amount)),
        'note': note,
        'period': period,
        'store_name': '',
    }


# Opening balance entry - use earliest date in the data
entries.append(make_entry(date(2023, 10, 1), 'BALANCE', opening_balance,
                          note='Opening Balance', period=''))

# Track current date and current month/year for rows without dates
last_date = None
current_year = 2024  # Starting year context from row 1
current_month = None

# Month name mapping (Korean)
MONTH_MAP = {
    '1월': 1, '2월': 2, '3월': 3, '4월': 4, '5월': 5, '6월': 6,
    '7월': 7, '8월': 8, '9월': 9, '10월': 10, '11월': 11, '12월': 12,
}

# Process rows 3 to end of real data
for row_num in range(3, ws.max_row + 1):
    date_val = ws.cell(row_num, 1).value
    income_val = ws.cell(row_num, 2).value
    expense_val = ws.cell(row_num, 3).value
    note_val = ws.cell(row_num, 4).value
    balance_val = ws.cell(row_num, 5).value

    # Clean up note
    note = str(note_val).strip() if note_val else ''
    if note in ('None', ''):
        note = ''

    # Skip rows where both income and expense are 0/None/empty AND note is empty
    income = income_val if income_val and income_val != 0 else 0
    expense = expense_val if expense_val and expense_val != 0 else 0

    if income == 0 and expense == 0:
        # Still process date changes for month markers
        if date_val:
            date_str = str(date_val).strip()
            if date_str in MONTH_MAP:
                current_month = MONTH_MAP[date_str]
                # Handle year rollover
                if current_month == 1 and last_date and last_date.month >= 10:
                    current_year += 1
            elif '11월' in date_str:
                current_month = 11
        continue

    # Determine the date for this row
    row_date = None
    if isinstance(date_val, datetime):
        row_date = date_val.date() if hasattr(date_val, 'date') else date_val
        current_year = row_date.year
        current_month = row_date.month
    elif isinstance(date_val, date) and not isinstance(date_val, datetime):
        row_date = date_val
        current_year = row_date.year
        current_month = row_date.month
    elif date_val:
        date_str = str(date_val).strip()
        if date_str in MONTH_MAP:
            current_month = MONTH_MAP[date_str]
            # Handle year rollover
            if current_month == 1 and last_date and last_date.month >= 10:
                current_year += 1
            row_date = date(current_year, current_month, 1)
        elif '11월' in date_str:
            current_month = 11
            row_date = date(current_year, current_month, 1)
        elif date_str == ' ':
            row_date = last_date
        else:
            # Try as integer (day only) - e.g. rows 122-131: 4, 5, 17, 27, 28
            try:
                day = int(date_str)
                if current_month and 1 <= day <= 31:
                    row_date = date(current_year, current_month, day)
            except ValueError:
                pass

    if row_date is None:
        if last_date:
            row_date = last_date
        else:
            print(f"  WARNING: Row {row_num} has no date context, skipping: {note}")
            continue

    last_date = row_date

    # Determine period based on date
    y = row_date.year
    m = row_date.month
    if m >= 10:
        period = f"{y}-Oct"
    elif m >= 4:
        period = f"{y}-Apr"
    else:
        period = f"{y - 1}-Oct"

    # Determine transaction type
    if income > 0 and expense > 0:
        # Both income and expense on same row (e.g. row 116: 여표 5000000/5000000)
        entries.append(make_entry(row_date, 'EXCHANGE', income,
                                  note=f'{note} (입금)', period=period))
        entries.append(make_entry(row_date, 'EXPENSE', expense,
                                  note=f'{note} (지출)', period=period))
        continue
    elif income > 0:
        typ = 'EXCHANGE'  # ChCh KRW income is exchange/inflow
    else:
        typ = 'EXPENSE'

    amount = income if income > 0 else expense
    if amount == 0:
        continue

    entries.append(make_entry(row_date, typ, amount, note=note, period=period))

# ============================================================
# Step 3: Create CQTransaction records
# ============================================================
print(f"\nTotal entries to import: {len(entries)}")
print(f"\nEntries preview:")
for i, e in enumerate(entries):
    direction = '+' if e['transaction_type'] in ('BALANCE', 'EXCHANGE', 'TRANSFER', 'COLLECTION') else '-'
    print(f"  {i+1:3d}. {e['date']} {e['transaction_type']:10s} {direction}{e['amount']:>15,.0f}  {e['note'][:60]}")

created = 0
for e in entries:
    CQTransaction.objects.create(
        organization=org,
        date=e['date'],
        store_name=e['store_name'],
        transaction_type=e['transaction_type'],
        person=PERSON,
        amount=e['amount'],
        account_type=ACCT,
        note=e['note'],
        period=e['period'],
    )
    created += 1

print(f"\nCreated {created} CQ Transactions for ChCh (KRW) account")

# ============================================================
# Step 4: Verify running balance
# ============================================================
INFLOW = ('COLLECTION', 'BALANCE', 'TRANSFER', 'EXCHANGE')
qs = CQTransaction.objects.filter(person=PERSON).order_by('date', 'id')
balance = Decimal('0')
for t in qs:
    if t.transaction_type in INFLOW:
        balance += t.amount
    else:
        balance -= t.amount

print(f"\nFinal ChCh KRW balance: {balance:,.0f} KRW")
print(f"Expected from Excel:    8,942,503 KRW")
diff = balance - Decimal('8942503')
print(f"Match: {'YES' if diff == 0 else 'NO - DIFF: ' + str(diff)}")

# Print summary by type
from django.db.models import Sum, Count
summary = qs.values('transaction_type').annotate(
    total=Sum('amount'), cnt=Count('id')
).order_by('transaction_type')
print(f"\nSummary by type:")
for s in summary:
    print(f"  {s['transaction_type']:12s}: {s['cnt']:3d} records, total = {s['total']:>15,.0f} KRW")
